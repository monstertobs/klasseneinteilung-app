"""
Klasseneinteilung App - Intelligente Klasseneinteilung für 5. Klassen

Version: 0.1.52
Author: Tobias Meier <admin(at)secutobs.com>
Date: 18. Juni 2026
License: Proprietary - All rights reserved

Description:
    Webanwendung zur automatisierten Erstellung von Klasseneinteilungen
    mit Berücksichtigung von Geschlechterverteilung, Schulweg, Elternwünschen,
    Schulformen, Religion, Spezialklassen und Inklusion.

Features:
    - Intelligenter Einteilungs-Algorithmus
    - Export (Excel, CSV, PDF)
    - Mehrere Spezialklassen (Sport, Musik, Theater)
    - Religion-Bündelung
    - IB Min/Max-Einschränkungen
    - Drag & Drop Vorschau-Modus
    - Konflikt-Erkennung und -Auflösung
    - DSGVO-konform
    - Sicherheits-Features (CSRF, Rate Limiting, sichere Sessions)
"""

__version__ = '0.1.52'
__author__ = 'Tobias Meier'
__email__ = 'admin(at)secutobs.com'

# GitHub Update-Konfiguration
GITHUB_REPO = 'monstertobs/klasseneinteilung-app'
GITHUB_RAW_BASE = f'https://raw.githubusercontent.com/{GITHUB_REPO}/main'
GITHUB_ZIP_URL = f'https://github.com/{GITHUB_REPO}/archive/refs/heads/main.zip'

# Konstanten
MAX_CLASS_SIZE = 25  # Maximale Anzahl Schüler pro Klasse
__copyright__ = 'Copyright © 2026 Tobias Meier'
__license__ = 'Proprietary'

import os
import sqlite3
import random
import csv
import io
import re
import secrets
import string
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_session import Session
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Lade .env Datei (für SECRET_KEY und andere Konfiguration)
try:
    from dotenv import load_dotenv
    load_dotenv()  # Lädt .env aus dem aktuellen Verzeichnis
except ImportError:
    pass  # python-dotenv nicht installiert (Production mit System-Variablen)

app = Flask(__name__)

# SECRET_KEY muss als Umgebungsvariable gesetzt sein (Sicherheit!)
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    raise RuntimeError(
        "KRITISCHER FEHLER: SECRET_KEY Umgebungsvariable nicht gesetzt!\n"
        "Generieren Sie einen sicheren Key mit:\n"
        "  python3 -c 'import secrets; print(secrets.token_hex(32))'\n"
        "Dann setzen Sie ihn in der .env Datei:\n"
        "  SECRET_KEY=<generierter-key>"
    )

# Session-Konfiguration für Filesystem-basierte Sessions (größere Daten möglich)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.path.join(os.path.dirname(__file__), 'flask_session')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True

# JSON und Encoding-Konfiguration
app.config['JSON_AS_ASCII'] = False  # Erlaubt Unicode in JSON (keine ASCII-Escape-Sequenzen)

# Sicherheits-Konfiguration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['WTF_CSRF_TIME_LIMIT'] = None  # CSRF-Token läuft mit Session ab
app.config['WTF_CSRF_SSL_STRICT'] = False  # Für Entwicklung

# CSRF-Schutz aktivieren
csrf = CSRFProtect(app)

# Session initialisieren
Session(app)

# Rate Limiting aktivieren (verhindert Brute-Force-Angriffe)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

IDLE_TIMEOUT_MINUTES = 15  # Automatischer Logout nach X Minuten Inaktivität
ONLINE_THRESHOLD_SECONDS = 5 * 60  # 5 Minuten = "online"

# KI-Assistent Konfiguration
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
KI_PROXY_URL = os.environ.get('KI_PROXY_URL', '')      # Für Stick: URL zu klassenwahl.de/api/ki-config
KI_PROXY_TOKEN = os.environ.get('KI_PROXY_TOKEN', '')  # Shared Secret für Proxy-Endpunkt

# In-Memory-Tracker: {user_id: datetime} — wird bei jedem Request aktualisiert
_online_users: dict = {}

@app.before_request
def check_idle_timeout():
    """Automatischer Logout nach IDLE_TIMEOUT_MINUTES Minuten Inaktivität.
    Prüft außerdem ob der Account noch freigeschaltet ist."""
    if 'user_id' not in session:
        return
    if request.endpoint in ('static', 'login', 'logout', 'register'):
        return

    now = datetime.now()

    # Online-Status aktualisieren
    _online_users[session['user_id']] = now

    # Freischaltungs-Check + Pending-Badge (immer setzen, auch bei Fehler)
    g.pending_users_count = 0
    try:
        db = get_main_db()
        if not session.get('is_admin'):
            row = db.execute('SELECT is_approved FROM users WHERE id = ?',
                             (session['user_id'],)).fetchone()
            if row and not dict(row).get('is_approved', 1):
                db.close()
                session.clear()
                flash('Ihr Konto wurde noch nicht freigeschaltet oder wurde gesperrt.', 'warning')
                return redirect(url_for('login'))
        else:
            row = db.execute('SELECT COUNT(*) FROM users WHERE is_approved = 0').fetchone()
            g.pending_users_count = row[0] if row else 0
        db.close()
    except Exception:
        pass  # Spalte existiert noch nicht (Migration läuft beim nächsten Start)

    last_activity = session.get('last_activity')
    if last_activity:
        idle_seconds = (now - datetime.fromisoformat(last_activity)).total_seconds()
        if idle_seconds > IDLE_TIMEOUT_MINUTES * 60:
            session.clear()
            flash('Sie wurden wegen Inaktivität automatisch abgemeldet.', 'warning')
            return redirect(url_for('login'))
    session['last_activity'] = now.isoformat()

# Response-Header für UTF-8 Encoding setzen
@app.after_request
def after_request(response):
    """Setzt UTF-8 Encoding und Security-Header für alle Responses"""
    if response.content_type and 'text/html' in response.content_type:
        response.content_type = 'text/html; charset=utf-8'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'no-referrer'
    return response

DATABASE = 'klasseneinteilung.db'
USER_DATA_DIR = 'user_data'


def get_user_db_path(user_id):
    """Pfad zur benutzerspezifischen Datenbank"""
    os.makedirs(USER_DATA_DIR, exist_ok=True)
    return os.path.join(USER_DATA_DIR, f'user_{user_id}.db')


def get_main_db():
    """Verbindung zur Haupt-Datenbank (nur users-Tabelle)"""
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.text_factory = str
    return db


def get_db():
    """Verbindung zur benutzerspezifischen Datenbank des eingeloggten Nutzers"""
    from flask import session
    user_id = session.get('user_id')
    if not user_id:
        raise RuntimeError('Kein Benutzer angemeldet')
    db_path = get_user_db_path(user_id)
    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    db.text_factory = str
    return db

def get_school_config():
    """Schul-Konfiguration aus DB laden (gibt immer ein dict zurück)"""
    defaults = {
        'ki_enabled': 0, 'gender_balance': 1, 'parent_wishes': 1,
        'religion_distribute': 1, 'religion_group': 0, 'religion_bundle': 0,
        'ib_min': 2, 'ib_max': 5, 'ib_class_size': 22, 'config_source': 'manual',
    }
    try:
        db = get_db()
        row = db.execute('SELECT * FROM school_config WHERE id = 1').fetchone()
        db.close()
        if row:
            return dict(row)
    except Exception:
        pass
    return defaults


def _ki_analyze_text(text):
    """Sendet Schulbeschreibung an Gemini (direkt oder via Proxy) und gibt Config-Dict zurück."""
    import json as _json
    import re as _re

    prompt = (
        "Du bist ein Schulverwaltungs-Assistent. Analysiere die folgende Schulbeschreibung "
        "und erstelle eine JSON-Konfiguration fuer einen Klasseneinteilungs-Algorithmus.\n\n"
        f"Schulbeschreibung: {text}\n\n"
        "Antworte NUR mit einem JSON-Objekt mit diesen Feldern:\n"
        "- gender_balance (true/false): Geschlechterbalance priorisieren\n"
        "- parent_wishes (true/false): Elternwuensche beruecksichtigen\n"
        "- religion_distribute (true/false): Religion gleichmaessig verteilen\n"
        "- religion_group (true/false): Gleichreligioese Schueler zusammenhalten\n"
        "- religion_bundle (true/false): Religion in einzelne Klassen buendeln\n"
        "- ib_min (Ganzzahl 0-5): Mindestanzahl IB-Schueler pro Klasse (0=keine IB)\n"
        "- ib_max (Ganzzahl 0-10): Maximum IB-Schueler pro Klasse (0=keine Beschraenkung)\n"
        "- ib_class_size (Ganzzahl 15-25): Klassengroesse fuer IB-Klassen\n\n"
        "Hinweis: religion_distribute, religion_group und religion_bundle schliessen "
        "sich gegenseitig aus – hoechstens eines davon true.\n"
        "Antworte ausschliesslich mit validem JSON, keine Erklaerungen."
    )

    try:
        import requests as _req

        if KI_PROXY_URL:
            # Portable Stick: Anfrage an klassenwahl.de weiterleiten
            resp = _req.post(
                KI_PROXY_URL,
                json={'text': text},
                headers={'X-KI-Token': KI_PROXY_TOKEN},
                timeout=30,
            )
            resp.raise_for_status()
            return resp.json()

        if GEMINI_API_KEY:
            # Direkt: Gemini REST API aufrufen
            url = (
                'https://generativelanguage.googleapis.com/v1beta/models/'
                f'gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}'
            )
            payload = {
                'contents': [{'parts': [{'text': prompt}]}],
                'generationConfig': {'temperature': 0.1},
            }
            resp = _req.post(url, json=payload, timeout=30)
            resp.raise_for_status()
            raw = resp.json()['candidates'][0]['content']['parts'][0]['text']
            # JSON aus der Antwort extrahieren (auch wenn Gemini Markdown zurückgibt)
            match = _re.search(r'\{.*\}', raw, _re.DOTALL)
            if match:
                return _json.loads(match.group())
            return _json.loads(raw)

        raise RuntimeError('KI nicht konfiguriert: weder GEMINI_API_KEY noch KI_PROXY_URL gesetzt.')

    except Exception as exc:
        # API-Key aus Fehlermeldung entfernen
        msg = str(exc).replace(GEMINI_API_KEY, '***') if GEMINI_API_KEY else str(exc)
        raise RuntimeError(f'KI-Analyse fehlgeschlagen: {msg}') from exc


def init_user_db(user_id):
    """Schema für die benutzerspezifische Datenbank anlegen / migrieren"""
    db_path = get_user_db_path(user_id)
    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    cursor = db.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firstname TEXT NOT NULL,
            lastname TEXT NOT NULL,
            gender TEXT,
            wohnort TEXT DEFAULT '',
            schulform TEXT DEFAULT '',
            religion TEXT DEFAULT '',
            sportlich INTEGER DEFAULT 0,
            special_needs TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            import_batch_id TEXT
        )
    ''')

    for col, coldef in [
        ('religion', 'TEXT DEFAULT ""'),
        ('sportlich', 'INTEGER DEFAULT 0'),
        ('import_batch_id', 'TEXT'),
        ('wohnort', 'TEXT DEFAULT ""'),
        ('schulform', 'TEXT DEFAULT ""'),
        ('sport_interesse', 'INTEGER DEFAULT 0'),
        ('musik_interesse', 'INTEGER DEFAULT 0'),
        ('theater_interesse', 'INTEGER DEFAULT 0'),
        ('ikl', 'INTEGER DEFAULT 0'),
    ]:
        try:
            cursor.execute(f'ALTER TABLE students ADD COLUMN {col} {coldef}')
        except sqlite3.OperationalError:
            pass

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS parent_wishes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            wish_type TEXT NOT NULL,
            related_student_id INTEGER,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students (id),
            FOREIGN KEY (related_student_id) REFERENCES students (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS class_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            data TEXT NOT NULL,
            username TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS school_config (
            id INTEGER PRIMARY KEY DEFAULT 1,
            ki_enabled INTEGER DEFAULT 0,
            gender_balance INTEGER DEFAULT 1,
            parent_wishes INTEGER DEFAULT 1,
            religion_distribute INTEGER DEFAULT 1,
            religion_group INTEGER DEFAULT 0,
            religion_bundle INTEGER DEFAULT 0,
            ib_min INTEGER DEFAULT 2,
            ib_max INTEGER DEFAULT 5,
            ib_class_size INTEGER DEFAULT 22,
            config_source TEXT DEFAULT 'manual',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('INSERT OR IGNORE INTO school_config (id) VALUES (1)')

    db.commit()
    db.close()


def _migrate_legacy_data():
    """Vorhandene Daten aus der monolithischen DB in Admin-User-DB übertragen (einmalig)"""
    marker = os.path.join(USER_DATA_DIR, '.migrated')
    if os.path.exists(marker):
        return

    main_db = get_main_db()
    cursor = main_db.cursor()

    # Prüfen ob alte Tabellen existieren
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='students'")
    if not cursor.fetchone():
        main_db.close()
        # Kein alter Datenbestand – trotzdem Marker setzen
        os.makedirs(USER_DATA_DIR, exist_ok=True)
        open(marker, 'w').close()
        return

    cursor.execute("SELECT id FROM users WHERE username = 'admin'")
    admin = cursor.fetchone()
    if not admin:
        main_db.close()
        return

    admin_id = admin['id']
    init_user_db(admin_id)
    user_db = sqlite3.connect(get_user_db_path(admin_id))
    user_db.row_factory = sqlite3.Row
    uc = user_db.cursor()

    # Schüler migrieren (IDs beibehalten wegen Elternwunsch-Referenzen)
    cursor.execute('SELECT * FROM students')
    for s in cursor.fetchall():
        sd = dict(s)
        cols = [c for c in sd.keys() if c != 'created_by']
        placeholders = ','.join('?' * len(cols))
        col_names = ','.join(cols)
        uc.execute(
            f'INSERT OR IGNORE INTO students ({col_names}) VALUES ({placeholders})',
            [sd[c] for c in cols]
        )

    # Elternwünsche migrieren
    cursor.execute('SELECT * FROM parent_wishes')
    for w in cursor.fetchall():
        wd = dict(w)
        cols = list(wd.keys())
        placeholders = ','.join('?' * len(cols))
        col_names = ','.join(cols)
        uc.execute(
            f'INSERT OR IGNORE INTO parent_wishes ({col_names}) VALUES ({placeholders})',
            [wd[c] for c in cols]
        )

    # Klasseneinteilungen migrieren
    cursor.execute('SELECT * FROM class_assignments')
    for a in cursor.fetchall():
        ad = dict(a)
        uc.execute(
            'INSERT OR IGNORE INTO class_assignments (id, name, data, username, created_at) VALUES (?,?,?,?,?)',
            (ad['id'], ad['name'], ad['data'], 'admin', ad.get('created_at', ''))
        )

    user_db.commit()
    user_db.close()
    main_db.close()

    os.makedirs(USER_DATA_DIR, exist_ok=True)
    open(marker, 'w').close()
    print("Migration: Bestehende Daten wurden in Admin-Datenbank übertragen.")


def init_db():
    """Haupt-Datenbank initialisieren (nur users-Tabelle) + User-DBs anlegen"""
    db = get_main_db()
    cursor = db.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_approved INTEGER NOT NULL DEFAULT 1,
            last_login TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Spalten-Migration für bestehende users-Tabellen
    for col, coldef in [
        ('is_approved', 'INTEGER NOT NULL DEFAULT 1'),
        ('last_login', 'TIMESTAMP'),
    ]:
        try:
            cursor.execute(f'ALTER TABLE users ADD COLUMN {col} {coldef}')
        except sqlite3.OperationalError:
            pass

    # Standard-Admin-User erstellen (falls nicht vorhanden)
    cursor.execute('SELECT * FROM users WHERE username = ?', ('admin',))
    if not cursor.fetchone():
        try:
            alphabet = string.ascii_letters + string.digits + string.punctuation
            initial_password = ''.join(secrets.choice(alphabet) for _ in range(16))
            password_hash = generate_password_hash(initial_password, method='pbkdf2:sha256')
            cursor.execute('INSERT INTO users (username, password_hash) VALUES (?, ?)',
                          ('admin', password_hash))
            db.commit()

            print("\n" + "="*70)
            print("WICHTIG: Neuer Admin-Account erstellt!")
            print("="*70)
            print(f"  Benutzername: admin")
            print(f"  Passwort:     {initial_password}")
            print("="*70)
            print("BITTE SOFORT NACH DEM ERSTEN LOGIN ÄNDERN!")
            print("Das Passwort wird nicht erneut angezeigt und ist nicht wiederherstellbar.")
            print("="*70 + "\n")

            with open('.initial_password', 'w') as f:
                f.write(initial_password)
        except Exception as e:
            print(f"Warning: Could not create admin user: {e}")

    db.commit()

    # User-DBs für alle vorhandenen Nutzer anlegen / migrieren
    cursor.execute('SELECT id FROM users')
    for row in cursor.fetchall():
        init_user_db(row['id'])

    db.close()

    # Einmalige Migration aus altem monolithischem Schema
    _migrate_legacy_data()

def login_required(f):
    """Decorator für Login-Pflicht"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Bitte melden Sie sich an.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator für Admin-Pflicht"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Bitte melden Sie sich an.', 'warning')
            return redirect(url_for('login'))
        if not session.get('is_admin'):
            flash('Zugriff verweigert. Nur Administratoren haben Zugang zu dieser Seite.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def validate_password(password):
    """
    Validiert Passwort-Sicherheit

    Anforderungen:
    - Mindestens 8 Zeichen
    - Mindestens einen Großbuchstaben
    - Mindestens einen Kleinbuchstaben
    - Mindestens eine Ziffer

    Returns: (is_valid, error_message)
    """
    if len(password) < 8:
        return False, "Passwort muss mindestens 8 Zeichen lang sein."

    if not re.search(r'[A-Z]', password):
        return False, "Passwort muss mindestens einen Großbuchstaben enthalten."

    if not re.search(r'[a-z]', password):
        return False, "Passwort muss mindestens einen Kleinbuchstaben enthalten."

    if not re.search(r'\d', password):
        return False, "Passwort muss mindestens eine Ziffer enthalten."

    return True, ""

def safe_int(value, default=0):
    """Wandelt einen Formularwert robust in int um.
    Nicht-numerische Eingaben (z.B. 'abc') führen zum default statt zu einem 500-Fehler."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

@app.route('/')
def index():
    """Startseite - weiterleiten zu Dashboard oder Login"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/version')
@login_required
def version():
    """Version und Autor-Informationen anzeigen"""
    return jsonify({
        'version': __version__,
        'author': __author__,
        'email': __email__,
        'copyright': __copyright__,
        'license': __license__,
        'release_date': 'February 21, 2026'
    })

@app.route('/about')
def about():
    """Über die Anwendung und Kontaktdaten"""
    return render_template('about.html')

@app.route('/algorithm')
def algorithm():
    """Algorithmus-Dokumentation - Wie funktioniert die Klasseneinteilung?"""
    return render_template('algorithm.html')

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")  # Brute-Force-Schutz: Max 10 Login-Versuche pro Minute
def login():
    """Login-Seite"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        # Input-Validierung
        if not username or not password:
            flash('Bitte füllen Sie alle Felder aus.', 'danger')
            return render_template('login.html')

        db = get_main_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
        user = cursor.fetchone()
        db.close()

        if user and check_password_hash(user['password_hash'], password):
            if not dict(user).get('is_approved', 1):
                flash('Ihr Konto wartet noch auf die Freischaltung durch den Administrator.', 'warning')
                return render_template('login.html', initial_password=None)
            init_user_db(user['id'])  # DB anlegen falls noch nicht vorhanden
            # last_login aktualisieren
            db2 = get_main_db()
            db2.execute('UPDATE users SET last_login = ? WHERE id = ?',
                        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
            db2.commit()
            db2.close()
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = (user['username'] == 'admin')
            session.permanent = True
            session.modified = True
            _online_users[user['id']] = datetime.now()
            if os.path.exists('.initial_password'):
                os.remove('.initial_password')
            flash('Erfolgreich angemeldet!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Ungültiger Benutzername oder Passwort.', 'danger')

    initial_password = None
    if os.path.exists('.initial_password'):
        with open('.initial_password', 'r') as f:
            initial_password = f.read().strip()

    return render_template('login.html', initial_password=initial_password)

@app.route('/logout')
def logout():
    """Logout"""
    session.clear()
    flash('Sie wurden abgemeldet.', 'info')
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def register():
    """Selbstregistrierung – neuer Benutzer mit eigener Datenbank"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')

        if not username or not password:
            flash('Benutzername und Passwort sind erforderlich.', 'danger')
        elif len(username) < 3 or len(username) > 30:
            flash('Benutzername muss 3–30 Zeichen lang sein.', 'danger')
        elif not re.match(r'^[a-zA-Z0-9_.-]+$', username):
            flash('Benutzername darf nur Buchstaben, Ziffern, _, . und - enthalten.', 'danger')
        elif username.lower() == 'admin':
            flash('Dieser Benutzername ist reserviert.', 'danger')
        elif password != password_confirm:
            flash('Passwörter stimmen nicht überein.', 'danger')
        else:
            is_valid, error_msg = validate_password(password)
            if not is_valid:
                flash(error_msg, 'danger')
            else:
                db = get_main_db()
                cursor = db.cursor()
                cursor.execute('SELECT COUNT(*) as count FROM users')
                if cursor.fetchone()['count'] >= 50:
                    db.close()
                    flash('Maximale Benutzeranzahl erreicht. Bitte wenden Sie sich an den Administrator.', 'danger')
                else:
                    cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
                    if cursor.fetchone():
                        db.close()
                        flash('Dieser Benutzername ist bereits vergeben.', 'danger')
                    else:
                        password_hash = generate_password_hash(password, method='pbkdf2:sha256')
                        # is_approved=0: neuer User wartet auf Admin-Freischaltung
                        cursor.execute(
                            'INSERT INTO users (username, password_hash, is_approved) VALUES (?, ?, 0)',
                            (username, password_hash)
                        )
                        db.commit()
                        new_id = cursor.lastrowid
                        db.close()
                        init_user_db(new_id)
                        return render_template('register.html', pending=True, registered_username=username)

    return render_template('register.html', pending=False)


@app.route('/dashboard')
@login_required
def dashboard():
    """Haupt-Dashboard"""
    import json as _json
    db = get_db()
    cursor = db.cursor()

    # Schüler-Statistiken
    cursor.execute('SELECT COUNT(*) as count FROM students')
    student_count = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students WHERE gender='m'")
    student_m = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students WHERE gender='w'")
    student_w = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students WHERE schulform='IB'")
    student_ib = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students WHERE ikl=1")
    student_ikl = cursor.fetchone()['count']

    # Wunsch-Statistiken
    cursor.execute('SELECT COUNT(*) as count FROM parent_wishes')
    wish_count = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM parent_wishes WHERE wish_type='together'")
    wish_together = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM parent_wishes WHERE wish_type='separated'")
    wish_separated = cursor.fetchone()['count']

    # Einteilungen
    cursor.execute('SELECT COUNT(*) as count FROM class_assignments')
    assignment_count = cursor.fetchone()['count']

    # Wunsch-Erfüllungsrate aus neuester Einteilung berechnen
    wish_rate = None
    wish_fulfilled = 0
    wish_total = 0
    latest_assignment_name = None

    cursor.execute('SELECT name, data FROM class_assignments ORDER BY created_at DESC LIMIT 1')
    latest = cursor.fetchone()
    if latest and wish_count > 0:
        latest_assignment_name = latest['name']
        try:
            proposal = _json.loads(latest['data'])
            student_class = {}
            for cls in proposal.get('classes', []):
                for s in cls.get('students', []):
                    student_class[s['id']] = cls['name']

            cursor.execute('SELECT student_id, related_student_id, wish_type FROM parent_wishes')
            wishes = cursor.fetchall()
            for w in wishes:
                if w['related_student_id'] not in student_class:
                    continue
                wish_total += 1
                s_cls = student_class.get(w['student_id'])
                r_cls = student_class.get(w['related_student_id'])
                if s_cls and r_cls:
                    if w['wish_type'] == 'together' and s_cls == r_cls:
                        wish_fulfilled += 1
                    elif w['wish_type'] == 'separated' and s_cls != r_cls:
                        wish_fulfilled += 1
            if wish_total > 0:
                wish_rate = round(wish_fulfilled / wish_total * 100)
        except Exception:
            pass

    db.close()

    # Wizard-Status prüfen
    wizard_active = session.get('wizard_active', False)
    wizard_step = session.get('wizard_step', 0)

    return render_template('dashboard.html',
                         student_count=student_count,
                         student_m=student_m,
                         student_w=student_w,
                         student_ib=student_ib,
                         student_ikl=student_ikl,
                         wish_count=wish_count,
                         wish_together=wish_together,
                         wish_separated=wish_separated,
                         assignment_count=assignment_count,
                         wish_rate=wish_rate,
                         wish_fulfilled=wish_fulfilled,
                         wish_total=wish_total,
                         latest_assignment_name=latest_assignment_name,
                         wizard_active=wizard_active,
                         wizard_step=wizard_step,
                         version=__version__,
                         author=__author__)

@app.route('/students')
@login_required
def students():
    """Schülerliste anzeigen"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM students ORDER BY lastname, firstname')
    students = cursor.fetchall()
    db.close()
    
    return render_template('students.html', students=students)

@app.route('/students/add', methods=['GET', 'POST'])
@login_required
def add_student():
    """Schüler hinzufügen"""
    if request.method == 'POST':
        firstname = request.form.get('firstname', '').strip()
        lastname = request.form.get('lastname', '').strip()
        gender = request.form.get('gender', '')
        wohnort = request.form.get('wohnort', '').strip()
        schulform = request.form.get('schulform', '')
        religion = request.form.get('religion', '')
        sportlich = 1 if 'sportlich' in request.form else 0
        sport_interesse = 1 if 'sport_interesse' in request.form else 0
        special_needs = request.form.get('special_needs', '').strip()
        ikl = 1 if 'ikl' in request.form else 0
        notes = request.form.get('notes', '').strip()

        if firstname and lastname:
            db = get_db()
            cursor = db.cursor()
            cursor.execute('''
                INSERT INTO students (firstname, lastname, gender, wohnort, schulform, religion, sportlich,
                                     sport_interesse, musik_interesse, theater_interesse, special_needs, ikl, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (firstname, lastname, gender, wohnort, schulform, religion, sportlich,
                  sport_interesse, 0, 0, special_needs, ikl, notes))
            db.commit()
            db.close()

            flash(f'Schüler {firstname} {lastname} wurde hinzugefügt.', 'success')
            return redirect(url_for('students'))
        else:
            flash('Vorname und Nachname sind erforderlich.', 'danger')

    return render_template('add_student.html')

@app.route('/students/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    """Schüler bearbeiten"""
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        firstname = request.form.get('firstname', '').strip()
        lastname = request.form.get('lastname', '').strip()
        gender = request.form.get('gender', '')
        wohnort = request.form.get('wohnort', '').strip()
        schulform = request.form.get('schulform', '')
        religion = request.form.get('religion', '')
        sportlich = 1 if 'sportlich' in request.form else 0
        sport_interesse = 1 if 'sport_interesse' in request.form else 0
        special_needs = request.form.get('special_needs', '').strip()
        ikl = 1 if 'ikl' in request.form else 0
        notes = request.form.get('notes', '').strip()

        if firstname and lastname:
            cursor.execute('''
                UPDATE students
                SET firstname = ?, lastname = ?, gender = ?, wohnort = ?, schulform = ?, religion = ?,
                    sportlich = ?, sport_interesse = ?, musik_interesse = ?, theater_interesse = ?,
                    special_needs = ?, ikl = ?, notes = ?
                WHERE id = ?
            ''', (firstname, lastname, gender, wohnort, schulform, religion, sportlich,
                  sport_interesse, 0, 0, special_needs, ikl, notes, student_id))
            db.commit()
            db.close()

            flash(f'Schüler {firstname} {lastname} wurde aktualisiert.', 'success')
            return redirect(url_for('students'))
        else:
            flash('Vorname und Nachname sind erforderlich.', 'danger')

    # Schüler-Daten laden
    cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
    student = cursor.fetchone()
    db.close()

    if not student:
        flash('Schüler nicht gefunden.', 'danger')
        return redirect(url_for('students'))

    return render_template('edit_student.html', student=student)

@app.route('/students/import', methods=['GET', 'POST'])
@login_required
def import_students():
    """Schüler aus CSV/Excel importieren"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Keine Datei ausgewählt.', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('Keine Datei ausgewählt.', 'danger')
            return redirect(request.url)

        if not file:
            flash('Ungültige Datei.', 'danger')
            return redirect(request.url)

        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        try:
            imported_count = 0
            errors = []
            duplicates = []

            # Eindeutige Batch-ID generieren
            import uuid
            batch_id = str(uuid.uuid4())[:8]

            # CSV-Import
            if file_ext == 'csv':
                stream = io.StringIO(file.stream.read().decode('utf-8-sig'), newline=None)
                csv_reader = csv.DictReader(stream, delimiter=';')

                # Fallback auf Komma, falls Semikolon nicht funktioniert
                if csv_reader.fieldnames and len(csv_reader.fieldnames) == 1:
                    stream.seek(0)
                    csv_reader = csv.DictReader(stream, delimiter=',')

                imported_count, errors, duplicates, wishes_created = process_import_data(csv_reader, batch_id)

            # Excel-Import
            elif file_ext in ['xlsx', 'xls']:
                if not EXCEL_SUPPORT:
                    flash('Excel-Unterstützung nicht installiert. Bitte installieren Sie openpyxl.', 'danger')
                    return redirect(request.url)

                workbook = load_workbook(filename=file)
                sheet = workbook.active

                # Header-Zeile lesen
                headers = [cell.value for cell in sheet[1]]

                # Daten als Dict-Liste aufbereiten
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Überspringe leere Zeilen
                        row_dict = dict(zip(headers, row))
                        data.append(row_dict)

                imported_count, errors, duplicates, wishes_created = process_import_data(data, batch_id)

            else:
                flash('Ungültiges Dateiformat. Bitte verwenden Sie CSV oder Excel (.xlsx).', 'danger')
                return redirect(request.url)

            # Zu Duplikats-Seite weiterleiten, wenn Duplikate gefunden wurden
            if duplicates:
                session['last_import_batch'] = batch_id
                if imported_count > 0:
                    flash(f'{imported_count} Schüler importiert.', 'success')
                if wishes_created > 0:
                    flash(f'{wishes_created} Freundewünsche importiert.', 'success')
                if errors:
                    session['import_errors'] = errors
                flash(f'{len(duplicates)} mögliche Duplikate gefunden — bitte prüfen.', 'warning')
                return redirect(url_for('check_duplicates'))

            # Ergebnis-Ansicht rendern (kein Redirect)
            import_results = {
                'filename': filename,
                'imported_count': imported_count,
                'errors': errors,
                'wishes_created': wishes_created,
            }
            return render_template('import_students.html',
                                   excel_support=EXCEL_SUPPORT,
                                   import_results=import_results)

        except Exception as e:
            flash(f'Fehler beim Import: {str(e)}', 'danger')
            return redirect(request.url)

    return render_template('import_students.html', excel_support=EXCEL_SUPPORT)

def process_import_data(data, batch_id):
    """Verarbeitet Import-Daten (CSV oder Excel)"""
    db = get_db()
    cursor = db.cursor()

    imported_count = 0
    errors = []
    duplicates = []
    wish_data = []  # Für Freundewünsche

    # Mögliche Spaltennamen (flexibel)
    field_mappings = {
        'firstname': ['vorname', 'firstname', 'first_name', 'first name', 'name'],
        'lastname': ['nachname', 'lastname', 'last_name', 'last name', 'familienname'],
        'gender': ['geschlecht', 'gender', 'sex'],
        'wohnort': ['wohnort', 'ort', 'adresse', 'address', 'schulweg', 'location', 'slr_wonadresse', 'slr_wohnadresse'],
        'schulform': ['schulform', 'schule', 'school_type', 'bildungsgang', 'eignung'],
        'religion': ['religionsunterricht', 'religion', 'konfession', 'wahlfach religion'],
        'sportlich': ['sportlich', 'athletic', 'sporty'],
        'sport_interesse': ['sportklasse'],  # Sportklassen-Hacken = Schüler soll in Sportklasse
        # Hinweis: 'Sportattest' wird absichtlich ignoriert (nur Attests-Status, kein Einfluss auf Sportklasse)
        'special_needs': ['förderbedarf', 'foerderbedarf', 'special_needs', 'special needs', 'sonderpädagogik'],
        'notes': ['notizen', 'notes', 'bemerkungen', 'anmerkungen', 'infos übergabe', 'infos uebergabe', 'sonstige / einwände', 'sonstige / einwaende', 'sonstige']
    }

    # Spezielle Spalten für Freundewünsche und IB/VM
    special_columns = {
        'freund_freundin': ['freund/ freundin', 'freund / freundin', 'freund/freundin', 'freund', 'freundin'],
        'auf_keinen_fall': ['auf keine fall mit kind…', 'auf keinen fall mit kind', 'auf keinen fall mit', 'auf keinen fall', 'nicht mit'],
        'ib_vm': ['ib / vm - s.liste', 'ib / vm', 'ib vm', 'ib/vm']
    }

    for idx, row in enumerate(data, start=1):
        try:
            # Flexibles Mapping
            student_data = {}
            row_lower = {k.lower().strip() if k else '': v for k, v in row.items()}

            # Vorname und Nachname sind Pflichtfelder
            firstname = None
            lastname = None
            notes_parts = []  # Für mehrere Notizen-Felder
            freund_value = None
            keinen_fall_value = None
            row_warnings = []  # Warnungen für diese Zeile (Schüler wird trotzdem importiert)

            # Standard-Felder mappen
            for db_field, possible_names in field_mappings.items():
                for possible_name in possible_names:
                    if possible_name in row_lower and row_lower[possible_name]:
                        value = row_lower[possible_name]

                        # Spezielle Behandlung für verschiedene Felder
                        if db_field == 'firstname':
                            firstname = str(value).strip()
                        elif db_field == 'lastname':
                            lastname = str(value).strip()
                        elif db_field == 'gender':
                            gender_map = {'m': 'm', 'männlich': 'm', 'male': 'm',
                                         'w': 'w', 'weiblich': 'w', 'female': 'w',
                                         'd': 'd', 'divers': 'd', 'diverse': 'd'}
                            mapped = gender_map.get(str(value).lower().strip())
                            if mapped:
                                student_data['gender'] = mapped
                            else:
                                student_data['gender'] = ''
                                row_warnings.append(f'Geschlecht "{str(value).strip()}" nicht erkannt — erwartet: m / w / d')
                        elif db_field == 'wohnort':
                            student_data['wohnort'] = str(value).strip()
                        elif db_field == 'schulform':
                            schulform_map = {'h': 'H', 'hauptschule': 'H', 'hs': 'H',
                                           'r': 'R', 'realschule': 'R', 'rs': 'R',
                                           'g': 'G', 'gymnasium': 'G', 'gym': 'G',
                                           'ib': 'IB', 'inklusiv': 'IB', 'inklusion': 'IB'}
                            mapped = schulform_map.get(str(value).lower().strip())
                            if mapped:
                                student_data['schulform'] = mapped
                            else:
                                student_data['schulform'] = str(value).strip()
                                row_warnings.append(f'Schulform "{str(value).strip()}" nicht erkannt — erwartet: H / R / G / IB')
                        elif db_field == 'religion':
                            rel_val = str(value).strip().lower()
                            if rel_val in ('katholisch', 'kath', 'rk', 'röm.-kath.', 'römisch-katholisch'):
                                student_data['religion'] = 'katholisch'
                            elif rel_val in ('evangelisch', 'ev', 'evang.', 'protestant', 'evangelisch-lutherisch'):
                                student_data['religion'] = 'evangelisch'
                            else:
                                student_data['religion'] = 'ethik'
                                if rel_val not in ('ethik', 'leer', ''):
                                    row_warnings.append(f'Religion "{str(value).strip()}" nicht erkannt — wird auf "Ethik" gesetzt')
                        elif db_field == 'sportlich':
                            sportlich_values = ['ja', 'yes', '1', 'true', 'x']
                            nein_values = ['nein', 'no', '0', 'false', '-', '']
                            val_low = str(value).lower().strip()
                            if val_low in sportlich_values:
                                student_data['sportlich'] = 1
                            elif val_low in nein_values:
                                student_data['sportlich'] = 0
                            else:
                                student_data['sportlich'] = 0
                                row_warnings.append(f'Sportlich "{str(value).strip()}" nicht erkannt — erwartet: ja / nein')
                        elif db_field == 'sport_interesse':
                            val_low = str(value).lower().strip()
                            # Erkennt auch "X" mit Kommentar (z.B. "X Sportattest liegt vor")
                            # und das ausgeschriebene Wort "Sportklasse"
                            is_sport = (
                                val_low in ('ja', 'yes', '1', 'true', 'x')
                                or val_low[:2] == 'x '
                                or 'sportklasse' in val_low
                            )
                            student_data['sport_interesse'] = 1 if is_sport else 0
                        elif db_field == 'special_needs':
                            sn_val = str(value).strip().lower()
                            known_sn = {'hoerschaedigung', 'sprache', 'sozial_emotional', 'lernen', 'sehen', 'kme'}
                            student_data['special_needs'] = str(value).strip()
                            if sn_val not in known_sn:
                                row_warnings.append(f'Förderbedarf "{str(value).strip()}" nicht erkannt — erwartet: hoerschaedigung / sprache / sozial_emotional / lernen / sehen / kme')
                        elif db_field == 'notes':
                            # Alle Notizen sammeln
                            notes_parts.append(f"{possible_name.title()}: {str(value).strip()}")
                        break

            # Konfession zusätzlich als Info in den Notizen festhalten
            # (nur wenn eine getrennte "Religionsunterricht"-Spalte existiert – dann ist
            #  "Religion"/"Konfession" die Mitgliedschaft, nicht der besuchte Unterricht)
            if 'religionsunterricht' in row_lower:
                konfession_raw = str(row_lower.get('religion') or row_lower.get('konfession') or '').strip()
                if konfession_raw and konfession_raw.lower() not in ('ethik', 'leer', 'keine', 'sonstige/keine'):
                    notes_parts.append(f"Konfession: {konfession_raw}")

            # Spezielle Spalten verarbeiten
            for special_field, possible_names in special_columns.items():
                for possible_name in possible_names:
                    if possible_name in row_lower and row_lower[possible_name]:
                        value = str(row_lower[possible_name]).strip()
                        if not value:
                            continue

                        if special_field == 'ib_vm':
                            # IB / VM Spalte: IB setzt schulform, VM wird als Notiz gespeichert
                            value_upper = value.upper()
                            if 'IB' in value_upper:
                                student_data['schulform'] = 'IB'
                            if 'VM' in value_upper:
                                notes_parts.append(f"VM: {value}")
                        elif special_field == 'freund_freundin':
                            freund_value = value
                        elif special_field == 'auf_keinen_fall':
                            keinen_fall_value = value
                        break

            # Notizen zusammenführen
            if notes_parts:
                student_data['notes'] = ' | '.join(notes_parts)

            # Eignung hat Vorrang: R oder H aus "Eignung"-Spalte überschreibt alles andere
            # (Elternwunsch Gymnasium z.B. in "Infos Übergabe" ändert die Schulform nicht)
            eignung_raw = row_lower.get('eignung', '')
            if eignung_raw:
                schulform_map_local = {'h': 'H', 'hauptschule': 'H', 'hs': 'H',
                                       'r': 'R', 'realschule': 'R', 'rs': 'R',
                                       'g': 'G', 'gymnasium': 'G', 'gym': 'G',
                                       'ib': 'IB'}
                eignung_mapped = schulform_map_local.get(str(eignung_raw).strip().lower())
                if eignung_mapped in ('R', 'H'):
                    current = student_data.get('schulform', '')
                    if current and current != eignung_mapped:
                        # Prüfen ob "Infos Übergabe" Gymnasium erwähnt
                        infos_raw = str(row_lower.get('infos übergabe', '') or row_lower.get('infos uebergabe', '')).lower()
                        if 'gymnasium' in infos_raw or ' g ' in f' {infos_raw} ':
                            row_warnings.append(f'Elternwunsch Gymnasium erkannt, aber Eignung ist {eignung_mapped} — Schüler bleibt bei {eignung_mapped}')
                        else:
                            row_warnings.append(f'Schulform "{current}" durch Eignung "{eignung_mapped}" überschrieben')
                    student_data['schulform'] = eignung_mapped

            # Pflichtfelder prüfen
            if not firstname or not lastname:
                partial = f"{firstname or ''} {lastname or ''}".strip() or '—'
                errors.append({'row': idx, 'name': partial, 'type': 'error', 'reason': 'Vorname oder Nachname fehlt'})
                continue

            # Duplikats-Check (gleicher Vor- und Nachname)
            cursor.execute('''
                SELECT id, firstname, lastname, gender, religion, import_batch_id
                FROM students
                WHERE LOWER(firstname) = LOWER(?) AND LOWER(lastname) = LOWER(?)
            ''', (firstname, lastname))

            existing = cursor.fetchone()
            if existing:
                duplicates.append({
                    'existing_id': existing['id'],
                    'name': f"{firstname} {lastname}",
                    'existing_is_import': existing['import_batch_id'] is not None
                })

            # In Datenbank einfügen (trotz Duplikat, wird später überprüft)
            cursor.execute('''
                INSERT INTO students (firstname, lastname, gender, wohnort, schulform, religion, sportlich, sport_interesse, special_needs, notes, import_batch_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                firstname,
                lastname,
                student_data.get('gender', ''),
                student_data.get('wohnort', ''),
                student_data.get('schulform', ''),
                student_data.get('religion', ''),
                student_data.get('sportlich', 0),
                student_data.get('sport_interesse', 0),
                student_data.get('special_needs', ''),
                student_data.get('notes', ''),
                batch_id
            ))

            student_id = cursor.lastrowid
            imported_count += 1

            # Feld-Warnungen für diese Zeile hinzufügen
            for w in row_warnings:
                errors.append({'row': idx, 'name': f'{firstname} {lastname}', 'type': 'warning', 'reason': w})

            # Freundewünsche für späteren Import speichern
            if freund_value:
                wish_data.append({
                    'student_id': student_id,
                    'student_name': f"{firstname} {lastname}",
                    'wish_type': 'together',
                    'related_names': freund_value
                })
            if keinen_fall_value:
                wish_data.append({
                    'student_id': student_id,
                    'student_name': f"{firstname} {lastname}",
                    'wish_type': 'separated',
                    'related_names': keinen_fall_value
                })

        except Exception as e:
            name = f"{firstname or ''} {lastname or ''}".strip() or '—'
            errors.append({'row': idx, 'name': name, 'type': 'error', 'reason': str(e)})

    db.commit()

    # Freundewünsche verarbeiten (nach dem Commit, damit alle Schüler in der DB sind)
    wishes_created = 0
    for wish in wish_data:
        try:
            # Namen aus der related_names Spalte extrahieren (kann mehrere Namen enthalten)
            related_names_raw = wish['related_names']
            # Namen können durch Komma, Semikolon oder "und" getrennt sein
            import re
            name_list = re.split(r'[,;]|\sund\s', related_names_raw)

            for related_name in name_list:
                related_name = related_name.strip()
                if not related_name:
                    continue

                # Namen splitten (Vorname Nachname oder Nachname, Vorname)
                name_parts = related_name.replace(',', ' ').split()
                if len(name_parts) < 2:
                    continue

                # Versuche verschiedene Kombinationen (Vorname Nachname und Nachname Vorname)
                possible_combinations = [
                    (name_parts[0], name_parts[1]),  # Vorname Nachname
                    (name_parts[1], name_parts[0])   # Nachname Vorname
                ]

                related_student_id = None
                for first, last in possible_combinations:
                    cursor.execute('''
                        SELECT id FROM students
                        WHERE LOWER(firstname) = LOWER(?) AND LOWER(lastname) = LOWER(?)
                    ''', (first, last))
                    result = cursor.fetchone()
                    if result:
                        related_student_id = result['id']
                        break

                if related_student_id and related_student_id != wish['student_id']:
                    # Prüfen ob Wunsch schon existiert (verhindert Duplikate)
                    cursor.execute('''
                        SELECT id FROM parent_wishes
                        WHERE student_id = ? AND related_student_id = ? AND wish_type = ?
                    ''', (wish['student_id'], related_student_id, wish['wish_type']))

                    if not cursor.fetchone():
                        cursor.execute('''
                            INSERT INTO parent_wishes (student_id, related_student_id, wish_type, description)
                            VALUES (?, ?, ?, ?)
                        ''', (wish['student_id'], related_student_id, wish['wish_type'],
                              f"Automatisch importiert aus: {wish['related_names']}"))
                        wishes_created += 1
                else:
                    # Name nicht gefunden - als Notiz hinzufügen
                    if related_student_id is None:
                        errors.append({'row': '—', 'name': wish['student_name'], 'type': 'warning', 'reason': f"Wunsch: '{related_name}' nicht gefunden"})

        except Exception as e:
            errors.append({'row': '—', 'name': wish['student_name'], 'type': 'warning', 'reason': f"Fehler bei Wunsch: {str(e)}"})

    db.commit()
    db.close()

    return imported_count, errors, duplicates, wishes_created

@app.route('/students/duplicates')
@login_required
def check_duplicates():
    """Duplikate überprüfen"""
    batch_id = session.get('last_import_batch')

    db = get_db()
    cursor = db.cursor()

    # Alle Schüler mit Duplikaten finden
    cursor.execute('''
        SELECT
            s1.id, s1.firstname, s1.lastname, s1.gender, s1.religion,
            s1.sportlich, s1.special_needs, s1.notes, s1.import_batch_id, s1.created_at,
            COUNT(*) OVER (PARTITION BY LOWER(s1.firstname), LOWER(s1.lastname)) as dup_count
        FROM students s1
        WHERE (
            SELECT COUNT(*)
            FROM students s2
            WHERE LOWER(s1.firstname) = LOWER(s2.firstname)
            AND LOWER(s1.lastname) = LOWER(s2.lastname)
        ) > 1
        ORDER BY LOWER(s1.lastname), LOWER(s1.firstname), s1.created_at
    ''')

    all_duplicates = cursor.fetchall()
    db.close()

    # Gruppiere nach Namen
    duplicates_by_name = {}
    for student in all_duplicates:
        key = f"{student['firstname']} {student['lastname']}".lower()
        if key not in duplicates_by_name:
            duplicates_by_name[key] = []
        duplicates_by_name[key].append(dict(student))

    return render_template('check_duplicates.html',
                         duplicates=duplicates_by_name,
                         batch_id=batch_id)

@app.route('/students/delete-duplicates', methods=['POST'])
@login_required
def delete_duplicates():
    """Ausgewählte Duplikate löschen"""
    student_ids = request.form.getlist('delete_ids[]')

    if not student_ids:
        flash('Keine Schüler zum Löschen ausgewählt.', 'warning')
        return redirect(url_for('check_duplicates'))

    db = get_db()
    cursor = db.cursor()

    for student_id in student_ids:
        # Verbundene Elternwünsche löschen
        cursor.execute('DELETE FROM parent_wishes WHERE student_id = ? OR related_student_id = ?',
                      (student_id, student_id))

        # Schüler löschen
        cursor.execute('DELETE FROM students WHERE id = ?', (student_id,))

    db.commit()
    db.close()

    flash(f'{len(student_ids)} Duplikat(e) erfolgreich gelöscht!', 'success')

    # Prüfen, ob noch Duplikate vorhanden
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT COUNT(*) as count
        FROM students s1
        WHERE (
            SELECT COUNT(*)
            FROM students s2
            WHERE LOWER(s1.firstname) = LOWER(s2.firstname)
            AND LOWER(s1.lastname) = LOWER(s2.lastname)
        ) > 1
    ''')
    remaining = cursor.fetchone()['count']
    db.close()

    if remaining > 0:
        return redirect(url_for('check_duplicates'))
    else:
        session.pop('last_import_batch', None)
        return redirect(url_for('students'))

@app.route('/students/delete/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    """Schüler löschen"""
    db = get_db()
    cursor = db.cursor()

    # Zuerst verbundene Elternwünsche löschen
    cursor.execute('DELETE FROM parent_wishes WHERE student_id = ? OR related_student_id = ?',
                  (student_id, student_id))

    # Dann den Schüler löschen
    cursor.execute('DELETE FROM students WHERE id = ?', (student_id,))
    db.commit()
    db.close()

    flash('Schüler wurde gelöscht.', 'success')
    return redirect(url_for('students'))

@app.route('/students/delete_multiple', methods=['POST'])
@login_required
def delete_multiple_students():
    """Mehrere ausgewählte Schüler löschen"""
    student_ids = request.form.getlist('student_ids')

    if not student_ids:
        flash('Keine Schüler ausgewählt.', 'warning')
        return redirect(url_for('students'))

    db = get_db()
    cursor = db.cursor()

    count = 0
    for student_id in student_ids:
        # Zuerst verbundene Elternwünsche löschen
        cursor.execute('DELETE FROM parent_wishes WHERE student_id = ? OR related_student_id = ?',
                      (student_id, student_id))

        # Dann den Schüler löschen
        cursor.execute('DELETE FROM students WHERE id = ?', (student_id,))
        count += 1

    db.commit()
    db.close()

    flash(f'{count} Schüler wurden gelöscht.', 'success')
    return redirect(url_for('students'))

@app.route('/students/delete_all', methods=['POST'])
@login_required
def delete_all_students():
    """Alle Schüler löschen"""
    db = get_db()
    cursor = db.cursor()

    # Zähle Schüler vor dem Löschen
    cursor.execute('SELECT COUNT(*) FROM students')
    count = cursor.fetchone()[0]

    if count == 0:
        flash('Keine Schüler vorhanden.', 'info')
        return redirect(url_for('students'))

    # Zuerst alle Elternwünsche löschen
    cursor.execute('DELETE FROM parent_wishes')

    # Dann alle Schüler löschen
    cursor.execute('DELETE FROM students')

    db.commit()
    db.close()

    flash(f'Alle {count} Schüler wurden gelöscht.', 'success')
    return redirect(url_for('students'))

@app.route('/students/generate_testdata', methods=['POST'])
@login_required
def generate_testdata():
    """Fiktive Testdaten generieren (nur für Testzwecke)"""

    # ── Namenslisten ──────────────────────────────────────────────
    VORNAMEN_M = [
        "Alexander", "Ben", "Daniel", "Emil", "Felix", "Jonas", "Leon", "Luca",
        "Maximilian", "Noah", "Paul", "Tim", "Tom", "Elias", "Finn", "Jan",
        "Luis", "Lukas", "Niklas", "Oscar", "Samuel", "Simon", "David", "Max",
        "Tobias", "Florian", "Moritz", "Julian", "Fabian", "Stefan",
    ]
    VORNAMEN_W = [
        "Anna", "Clara", "Emma", "Hannah", "Julia", "Laura", "Lea", "Lena",
        "Lisa", "Maria", "Mia", "Paula", "Sarah", "Sophie", "Charlotte", "Emily",
        "Emilia", "Johanna", "Lara", "Luisa", "Marie", "Nele", "Amelie", "Zoe",
        "Katharina", "Franziska", "Nicole", "Sandra", "Isabella", "Valentina",
    ]
    NACHNAMEN = [
        "Müller", "Schmidt", "Schneider", "Fischer", "Weber", "Meyer", "Wagner",
        "Becker", "Schulz", "Hoffmann", "Koch", "Bauer", "Richter", "Klein",
        "Wolf", "Schröder", "Neumann", "Braun", "Werner", "Schwarz",
        "Zimmermann", "Krüger", "Hartmann", "Lange", "Schmitt", "Krause",
        "Meier", "Lehmann", "Huber", "Mayer", "Herrmann", "König", "Walter",
        "Peters", "Lang", "Berger", "Winkler", "Frank", "Vogel", "Roth",
        "Beck", "Brandt", "Haas", "Schäfer", "Graf", "Fuchs", "Kaiser",
    ]
    ORTE = [
        "Musterstadt", "Musterstadt", "Musterstadt", "Musterstadt",
        "Nordviertel", "Nordviertel", "Südviertel", "Südviertel",
        "Westend", "Westend", "Kleinbach", "Kleinbach",
        "Großhausen", "Großhausen", "Feldkirchen", "Feldkirchen",
        "Bergheim", "Bergheim", "Seebach", "Waldorf",
        "Oberdorf", "Niederdorf", "Steinhausen", "Kirchdorf",
    ]
    RELIGIONEN = ["ethik", "ethik", "katholisch", "katholisch", "evangelisch", "evangelisch", ""]
    FOERDERBEDARFE = ["hoerschaedigung", "sprache", "sozial_emotional", "lernen", "sehen", "kme"]
    NOTIZEN = [
        "Sehr schüchtern, braucht Zeit",
        "Schulangst bekannt",
        "Fremdsprachig, Deutsch noch ausbaufähig",
        "Teilleistungsschwäche Lesen",
        "LRS-Verdacht",
        "Hochbegabung vermutet",
    ]

    try:
        anzahl = int(request.form.get('anzahl', 100))
        anzahl = max(10, min(300, anzahl))  # Sicherheitsklammer: 10–300
    except (ValueError, TypeError):
        anzahl = 100

    with_wishes = request.form.get('with_wishes') == '1'
    with_ib = request.form.get('with_ib') == '1'
    with_vm = request.form.get('with_vm') == '1'
    with_foerder = request.form.get('with_foerder') == '1'

    batch_id = f"TESTDATA-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    rng = random.Random()  # eigener RNG, beeinflusst nicht globalen Zustand

    db = get_db()
    cursor = db.cursor()

    # ── Duplikate im gleichen Batch vermeiden ─────────────────────
    used_names = set()

    def pick_name(geschlecht):
        for _ in range(300):
            vn = rng.choice(VORNAMEN_M if geschlecht == 'm' else VORNAMEN_W)
            nn = rng.choice(NACHNAMEN)
            key = (vn.lower(), nn.lower())
            if key not in used_names:
                used_names.add(key)
                return vn, nn
        return (rng.choice(VORNAMEN_M if geschlecht == 'm' else VORNAMEN_W),
                rng.choice(NACHNAMEN) + str(rng.randint(2, 9)))

    # ── IB / VM / Förderbedarf Indices bestimmen ──────────────────
    all_indices = list(range(anzahl))
    rng.shuffle(all_indices)

    ib_count = round(anzahl * 0.10) if with_ib else 0
    vm_count = min(12, round(anzahl * 0.12)) if with_vm else 0
    foerder_count = min(8, round(anzahl * 0.08)) if with_foerder else 0

    ib_set = set(all_indices[:ib_count])
    vm_set = set(all_indices[ib_count:ib_count + vm_count])
    foerder_set = set(all_indices[ib_count + vm_count:ib_count + vm_count + foerder_count])

    # ── Schüler erzeugen ──────────────────────────────────────────
    student_ids = []

    for i in range(anzahl):
        geschlecht = rng.choice(['m', 'm', 'w', 'w', 'd'])
        if geschlecht == 'd':
            geschlecht = rng.choice(['m', 'w'])

        vorname, nachname = pick_name(geschlecht)
        wohnort = rng.choice(ORTE)
        religion = rng.choice(RELIGIONEN)
        sportlich = 1 if rng.random() < 0.22 else 0
        sport_interesse = 1 if rng.random() < 0.15 else 0

        if i in ib_set:
            schulform = 'IB'
        else:
            r = rng.random()
            schulform = 'H' if r < 0.25 else ('R' if r < 0.62 else 'G')

        special_needs = ''
        notes_parts = []
        if i in foerder_set:
            special_needs = rng.choice(FOERDERBEDARFE)
        if i in vm_set:
            notes_parts.append('VM: Vorbeugende Maßnahme')
        if rng.random() < 0.08:
            notes_parts.append(rng.choice(NOTIZEN))

        notes = ' | '.join(notes_parts)

        cursor.execute('''
            INSERT INTO students
                (firstname, lastname, gender, wohnort, schulform, religion,
                 sportlich, sport_interesse, special_needs, notes,
                 import_batch_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (vorname, nachname, geschlecht, wohnort, schulform, religion,
              sportlich, sport_interesse, special_needs, notes,
              batch_id))

        student_ids.append(cursor.lastrowid)

    db.commit()

    # ── Elternwünsche ─────────────────────────────────────────────
    wishes_created = 0
    if with_wishes and len(student_ids) >= 4:
        together_count = round(len(student_ids) * 0.25)
        separated_count = round(len(student_ids) * 0.10)
        used_in_wish = set()

        shuffled = student_ids[:]
        rng.shuffle(shuffled)

        # Zusammen-Wünsche (Paare, bidirektional)
        i = 0
        pairs = 0
        while pairs < together_count // 2 and i + 1 < len(shuffled):
            a, b = shuffled[i], shuffled[i + 1]
            if a not in used_in_wish and b not in used_in_wish:
                cursor.execute(
                    'INSERT INTO parent_wishes (student_id, wish_type, related_student_id, description) VALUES (?, ?, ?, ?)',
                    (a, 'together', b, 'Testdaten – Freundschaftswunsch'))
                cursor.execute(
                    'INSERT INTO parent_wishes (student_id, wish_type, related_student_id, description) VALUES (?, ?, ?, ?)',
                    (b, 'together', a, 'Testdaten – Freundschaftswunsch'))
                used_in_wish.add(a)
                used_in_wish.add(b)
                wishes_created += 2
                pairs += 1
            i += 2

        # Getrennt-Wünsche (einseitig)
        remaining = [sid for sid in shuffled if sid not in used_in_wish]
        rng.shuffle(remaining)
        for j in range(0, min(separated_count * 2, len(remaining) - 1), 2):
            a, b = remaining[j], remaining[j + 1]
            cursor.execute(
                'INSERT INTO parent_wishes (student_id, wish_type, related_student_id, description) VALUES (?, ?, ?, ?)',
                (a, 'separated', b, 'Testdaten – Trennungswunsch'))
            wishes_created += 1

        db.commit()

    db.close()

    msg = f'{anzahl} fiktive Testschüler wurden generiert (Batch: {batch_id})'
    if wishes_created:
        msg += f', {wishes_created} Elternwünsche erstellt'
    msg += '. Testdaten sind an der Markierung «Importiert» erkennbar.'
    flash(msg, 'success')
    return redirect(url_for('students'))


@app.route('/students/delete_testdata', methods=['POST'])
@login_required
def delete_testdata():
    """Alle Testdaten-Schüler (import_batch_id beginnt mit 'TESTDATA-') löschen"""
    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT id FROM students WHERE import_batch_id LIKE 'TESTDATA-%'")
    ids = [row[0] for row in cursor.fetchall()]

    if not ids:
        flash('Keine Testdaten vorhanden.', 'info')
        db.close()
        return redirect(url_for('students'))

    placeholders = ','.join('?' * len(ids))
    cursor.execute(f'DELETE FROM parent_wishes WHERE student_id IN ({placeholders})', ids)
    cursor.execute(f'DELETE FROM parent_wishes WHERE related_student_id IN ({placeholders})', ids)
    cursor.execute("DELETE FROM students WHERE import_batch_id LIKE ?", ('TESTDATA-%',))
    db.commit()
    db.close()

    flash(f'{len(ids)} Testschüler und alle zugehörigen Elternwünsche wurden gelöscht.', 'success')
    return redirect(url_for('students'))


@app.route('/wishes')
@login_required
def wishes():
    """Elternwünsche anzeigen"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT pw.*, 
               s1.firstname || ' ' || s1.lastname as student_name,
               s2.firstname || ' ' || s2.lastname as related_student_name
        FROM parent_wishes pw
        JOIN students s1 ON pw.student_id = s1.id
        LEFT JOIN students s2 ON pw.related_student_id = s2.id
        ORDER BY s1.lastname, s1.firstname
    ''')
    wishes = cursor.fetchall()
    db.close()
    
    return render_template('wishes.html', wishes=wishes)

@app.route('/wishes/add', methods=['GET', 'POST'])
@login_required
def add_wish():
    """Elternwunsch hinzufügen"""
    db = get_db()
    cursor = db.cursor()
    
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        wish_type = request.form.get('wish_type')
        related_student_id = request.form.get('related_student_id') or None
        description = request.form.get('description', '').strip()
        
        if student_id and wish_type:
            cursor.execute('''
                INSERT INTO parent_wishes (student_id, wish_type, related_student_id, description)
                VALUES (?, ?, ?, ?)
            ''', (student_id, wish_type, related_student_id, description))
            db.commit()
            flash('Elternwunsch wurde hinzugefügt.', 'success')
            db.close()
            return redirect(url_for('wishes'))
        else:
            flash('Schüler und Wunsch-Typ sind erforderlich.', 'danger')
    
    # Schülerliste für Dropdown
    cursor.execute('SELECT * FROM students ORDER BY lastname, firstname')
    students = cursor.fetchall()
    db.close()
    
    return render_template('add_wish.html', students=students)

@app.route('/wishes/edit/<int:wish_id>', methods=['GET', 'POST'])
@login_required
def edit_wish(wish_id):
    """Elternwunsch bearbeiten"""
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        wish_type = request.form.get('wish_type')
        related_student_id = request.form.get('related_student_id') or None
        description = request.form.get('description', '').strip()

        if student_id and wish_type:
            cursor.execute('''
                UPDATE parent_wishes
                SET student_id = ?, wish_type = ?, related_student_id = ?, description = ?
                WHERE id = ?
            ''', (student_id, wish_type, related_student_id, description, wish_id))
            db.commit()
            flash('Elternwunsch wurde aktualisiert.', 'success')
            db.close()
            return redirect(url_for('wishes'))
        else:
            flash('Schüler und Wunsch-Typ sind erforderlich.', 'danger')

    # Wunsch laden
    cursor.execute('SELECT * FROM parent_wishes WHERE id = ?', (wish_id,))
    wish = cursor.fetchone()

    # Schülerliste für Dropdown
    cursor.execute('SELECT * FROM students ORDER BY lastname, firstname')
    students = cursor.fetchall()
    db.close()

    if not wish:
        flash('Elternwunsch nicht gefunden.', 'danger')
        return redirect(url_for('wishes'))

    return render_template('edit_wish.html', wish=wish, students=students)

@app.route('/wishes/delete/<int:wish_id>', methods=['POST'])
@login_required
def delete_wish(wish_id):
    """Elternwunsch löschen"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('DELETE FROM parent_wishes WHERE id = ?', (wish_id,))
    db.commit()
    db.close()

    flash('Elternwunsch wurde gelöscht.', 'success')
    return redirect(url_for('wishes'))

@app.route('/generate', methods=['GET', 'POST'])
@login_required
def generate():
    """Klasseneinteilungen generieren"""
    db = get_db()
    cursor = db.cursor()

    # Alle Schüler abrufen
    cursor.execute('SELECT * FROM students ORDER BY lastname, firstname')
    students = cursor.fetchall()

    # Alle Elternwünsche abrufen
    cursor.execute('SELECT * FROM parent_wishes')
    wishes = cursor.fetchall()

    # Gespeicherte Einteilungen für Basis-Auswahl laden
    cursor.execute('SELECT id, name, created_at FROM class_assignments ORDER BY created_at DESC')
    existing_assignments = cursor.fetchall()

    db.close()

    if len(students) == 0:
        flash('Keine Schüler vorhanden. Bitte fügen Sie zuerst Schüler hinzu.', 'warning')
        return redirect(url_for('students'))

    # Schulamt-Vorgabe: max. Klassenanzahl bei 25 Schüler/Klasse
    schulamt_max_classes = max(1, (len(students) + MAX_CLASS_SIZE - 1) // MAX_CLASS_SIZE)
    num_classes = schulamt_max_classes

    # Optionen aus POST oder Standardwerte
    if request.method == 'POST':
        # IB Min/Max
        ib_min = safe_int(request.form.get('ib_min', '0'), 0)
        ib_max = safe_int(request.form.get('ib_max', '0'), 0)

        # Validierung: min <= max
        if ib_min > ib_max and ib_max > 0:
            flash('IB-Minimum darf nicht größer als Maximum sein.', 'danger')
            ib_min = 0
            ib_max = 0

        # Spezialklassen
        specialized_classes = {}
        specialized_classes['sport'] = safe_int(request.form.get('specialized_sport_count', '0'), 0)

        # Freie Spezialklasse
        custom_count = safe_int(request.form.get('specialized_custom_count', '0'), 0)
        custom_name = request.form.get('specialized_custom_name', '').strip()
        if custom_count > 0 and custom_name:
            specialized_classes['custom'] = custom_count
            specialized_classes['custom_name'] = custom_name
        else:
            specialized_classes['custom'] = 0
            specialized_classes['custom_name'] = ''

        ib_class_size = safe_int(request.form.get('ib_class_size', '22'), 22)

        # Klassenanzahl-Override aus Formular (früh lesen, damit Checks korrekt sind)
        num_classes_form = safe_int(request.form.get('num_classes', schulamt_max_classes), schulamt_max_classes)
        num_classes_form = max(1, min(20, num_classes_form))
        num_classes = num_classes_form

        # Pre-Generation Check: Genug IB-Schüler?
        if ib_max > 0:
            ib_students = [s for s in students if s['schulform'] == 'IB']
            if len(ib_students) > num_classes * ib_max:
                flash(f'⚠️ Zu viele IB-Schüler ({len(ib_students)}) für {num_classes} Klassen mit max {ib_max} pro Klasse', 'warning')

        avg_class_size = len(students) / num_classes
        if avg_class_size > MAX_CLASS_SIZE:
            flash(f'⚠️ {num_classes} Klassen für {len(students)} Schüler ergibt Ø {avg_class_size:.1f} Schüler/Klasse (über dem Schulamt-Richtwert von {MAX_CLASS_SIZE}).', 'warning')

        options = {
            'gender_balance': 'gender_balance' in request.form,
            'parent_wishes': 'parent_wishes' in request.form,
            'religion_distribute': 'religion_distribute' in request.form,
            'religion_group': 'religion_group' in request.form,
            'religion_bundle': 'religion_bundle' in request.form,
            'sportklasse': 'sportklasse' in request.form,
            'specialized_classes': specialized_classes,
            'ib_min': ib_min,
            'ib_max': ib_max,
            'ib_class_size': ib_class_size,
            'num_classes': num_classes_form,
        }
    else:
        sc = get_school_config()
        options = {
            'gender_balance':      bool(sc.get('gender_balance', 1)),
            'parent_wishes':       bool(sc.get('parent_wishes', 1)),
            'religion_distribute': bool(sc.get('religion_distribute', 1)),
            'religion_group':      bool(sc.get('religion_group', 0)),
            'religion_bundle':     bool(sc.get('religion_bundle', 0)),
            'sportklasse': False,
            'specialized_classes': {'sport': 0, 'custom': 0, 'custom_name': ''},
            'ib_min':        sc.get('ib_min', 2),
            'ib_max':        sc.get('ib_max', 5),
            'ib_class_size': sc.get('ib_class_size', 22),
            'num_classes':   schulamt_max_classes,
        }

    # Bei GET: Spezialklassen-Mindestanzahl sicherstellen (Nutzer hat noch nicht explizit gewählt)
    if request.method == 'GET':
        sport_count_opt = options.get('specialized_classes', {}).get('sport', 0)
        if sport_count_opt > 0:
            non_sport_student_count = sum(1 for s in students if not s['sport_interesse'])
            non_sport_classes_min = max(1, (non_sport_student_count + MAX_CLASS_SIZE - 1) // MAX_CLASS_SIZE)
            num_classes = max(num_classes, sport_count_opt + non_sport_classes_min)

    # Basis-Einteilung laden (falls angegeben)
    base_assignment = None
    base_assignment_id = request.form.get('base_assignment_id', '') if request.method == 'POST' else ''
    if base_assignment_id:
        try:
            db2 = get_db()
            row = db2.execute('SELECT data FROM class_assignments WHERE id = ?', (int(base_assignment_id),)).fetchone()
            db2.close()
            if row:
                base_assignment = json.loads(row['data'])
        except Exception:
            pass

    # 3 verschiedene Einteilungen generieren
    proposals = []
    for i in range(3):
        proposal = generate_class_assignment(students, wishes, num_classes, i, options, base_assignment=base_assignment)
        proposals.append(proposal)

    # ── Vergleichsmetriken pro Vorschlag berechnen ────────────────
    import math as _math
    for proposal in proposals:
        # Schüler → Klassen-Mapping aufbauen
        student_class_map = {}
        for cls in proposal['classes']:
            for s in cls['students']:
                student_class_map[s['id']] = cls['number']

        # Wunscherfüllungsrate
        wish_total = wish_fulfilled = 0
        for w in wishes:
            rid = w['related_student_id']
            if not rid or rid not in student_class_map:
                continue
            wish_total += 1
            sc = student_class_map.get(w['student_id'])
            rc = student_class_map.get(rid)
            if sc and rc:
                if w['wish_type'] == 'together' and sc == rc:
                    wish_fulfilled += 1
                elif w['wish_type'] == 'separated' and sc != rc:
                    wish_fulfilled += 1
        proposal['statistics']['wish_fulfilled'] = wish_fulfilled
        proposal['statistics']['wish_total'] = wish_total
        proposal['statistics']['wish_rate'] = (
            round(wish_fulfilled / wish_total * 100) if wish_total > 0 else None
        )

        # Geschlechterbalance-Score (100 = perfekt ausgeglichen)
        ratios = []
        for cls in proposal['classes']:
            m_cnt = cls['gender_count']['m']
            w_cnt = cls['gender_count']['w']
            total_gend = m_cnt + w_cnt
            if total_gend > 0:
                ratios.append(m_cnt / total_gend)
        if ratios:
            mean_r = sum(ratios) / len(ratios)
            stddev = _math.sqrt(sum((r - mean_r) ** 2 for r in ratios) / len(ratios))
            proposal['statistics']['gender_balance_score'] = max(0, round((1 - stddev * 6) * 100))
        else:
            proposal['statistics']['gender_balance_score'] = 100

        # Schüler-IDs pro Klasse als Liste für JS-Vergleich
        proposal['statistics']['student_class_map'] = student_class_map

    # Proposals in Session speichern für Export
    session['last_proposals'] = proposals

    return render_template('generate.html', proposals=proposals, num_classes=num_classes, options=options,
                           schulamt_max_classes=schulamt_max_classes,
                           existing_assignments=existing_assignments, selected_base_id=base_assignment_id,
                           wish_count=len(wishes))

def extract_city_from_wohnort(wohnort):
    """Extrahiere Stadt (PLZ + Stadtname) aus vollständiger Adresse"""
    import re
    if not wohnort or not wohnort.strip():
        return None
    match = re.search(r'(\d{5})\s+([^,]+)', wohnort)
    if match:
        plz = match.group(1)
        stadt = match.group(2).strip()
        # Entferne Klammern wie "(Taunus)"
        stadt = re.sub(r'\s*\([^)]*\)', '', stadt).strip()
        return f"{plz} {stadt}"
    return None

def extract_plz_from_wohnort(wohnort):
    """Extrahiere nur die PLZ aus vollständiger Adresse"""
    import re
    if not wohnort or not wohnort.strip():
        return None
    match = re.search(r'(\d{5})', wohnort)
    if match:
        return match.group(1)
    return None

def optimize_assignment_wishes(classes, wish_dict, max_rounds=60, options=None, pinned_student_ids=None, specialized_mapping=None):
    """
    Post-Processing: Verbessert Wunsch-Erfüllungsrate durch iterativen Schüler-Tausch.
    - Tauscht bevorzugt gleichgeschlechtliche Schüler (Geschlechterbalance bleibt erhalten)
    - Verschiebt Schüler wenn Zielklasse noch Platz hat
    - Wiederholt bis keine Verbesserung mehr möglich (max. max_rounds Durchläufe)
    - Gepinnte Schüler (aus Basis-Einteilung) werden nicht bewegt
    - Sport-Schüler bleiben in Sportklassen; Nicht-Sport-Schüler kommen nicht in Sportklassen
    """
    num_classes = len(classes)
    if options is None:
        options = {}
    if pinned_student_ids is None:
        pinned_student_ids = set()
    if specialized_mapping is None:
        specialized_mapping = {}
    ib_min = options.get('ib_min', 0)
    ib_max = options.get('ib_max', 0)

    sport_class_indices = {ci for ci, t in specialized_mapping.items() if t == 'sport'}

    def sport_move_allowed(student, from_idx, to_idx):
        if not sport_class_indices:
            return True
        if student.get('sport_interesse'):
            # Sport-Schüler nicht aus Sportklasse herausbewegen
            if from_idx in sport_class_indices and to_idx not in sport_class_indices:
                return False
        else:
            # Nicht-Sport-Schüler nicht in Sportklasse hineinschieben
            if to_idx in sport_class_indices:
                return False
        return True

    def ib_count_for(cls):
        return sum(1 for s in cls if s.get('schulform') == 'IB')

    def ib_move_allowed(student, from_cls, to_cls):
        """Prüft ob Verschieben eines IB-Schülers die Constraints verletzt."""
        if student.get('schulform') != 'IB' or ib_min == 0:
            return True
        # Quellklasse darf nicht unter ib_min fallen (außer sie hat bereits zu wenig)
        from_ib = ib_count_for(from_cls)
        if from_ib - 1 > 0 and from_ib - 1 < ib_min:
            return False  # würde Einzelkämpfer in Quellklasse erzeugen
        # Zielklasse darf nicht über ib_max steigen
        if ib_max > 0 and ib_count_for(to_cls) + 1 > ib_max:
            return False
        return True

    # student_id → Klassen-Index
    student_to_class = {}
    for i, cls in enumerate(classes):
        for s in cls:
            student_to_class[s['id']] = i

    # Schneller Zugriff auf Schüler-Objekte
    all_students = {s['id']: s for cls in classes for s in cls}

    def total_wish_score():
        score = 0
        for sid, wlist in wish_dict.items():
            if sid not in student_to_class:
                continue
            for w in wlist:
                rid = w['related_student_id']
                if not rid or rid not in student_to_class:
                    continue
                same = student_to_class[sid] == student_to_class[rid]
                if w['wish_type'] == 'together' and same:
                    score += 1
                elif w['wish_type'] == 'separated' and not same:
                    score += 1
        return score

    current_score = total_wish_score()

    for _ in range(max_rounds):
        improved = False

        # --- Unerfüllte Zusammen-Wünsche ---
        for sid, wlist in list(wish_dict.items()):
            for w in wlist:
                if w['wish_type'] != 'together':
                    continue
                rid = w['related_student_id']
                if not rid:
                    continue
                if sid not in student_to_class or rid not in student_to_class:
                    continue

                class_a = student_to_class[sid]
                class_b = student_to_class[rid]
                if class_a == class_b:
                    continue

                s_obj = all_students.get(sid)
                if not s_obj:
                    continue
                if sid in pinned_student_ids:
                    continue
                gender = s_obj.get('gender', '')

                # Option 1: Verschiebe sid nach class_b wenn Platz vorhanden
                if (len(classes[class_b]) < MAX_CLASS_SIZE
                        and ib_move_allowed(s_obj, classes[class_a], classes[class_b])
                        and sport_move_allowed(s_obj, class_a, class_b)):
                    student_to_class[sid] = class_b
                    new_score = total_wish_score()
                    if new_score > current_score:
                        classes[class_a].remove(s_obj)
                        classes[class_b].append(s_obj)
                        current_score = new_score
                        improved = True
                        break
                    else:
                        student_to_class[sid] = class_a  # rückgängig

                # Option 2: Tausche sid mit gleichgeschlechtlichem Schüler aus class_b
                for candidate in list(classes[class_b]):
                    if candidate['id'] == rid:
                        continue
                    if candidate['id'] in pinned_student_ids:
                        continue
                    if candidate.get('gender', '') != gender:
                        continue
                    if not ib_move_allowed(s_obj, classes[class_a], classes[class_b]):
                        continue
                    if not ib_move_allowed(candidate, classes[class_b], classes[class_a]):
                        continue
                    if not sport_move_allowed(s_obj, class_a, class_b):
                        continue
                    if not sport_move_allowed(candidate, class_b, class_a):
                        continue
                    cid = candidate['id']
                    # In-place tauschen und testen
                    student_to_class[sid] = class_b
                    student_to_class[cid] = class_a
                    new_score = total_wish_score()
                    if new_score > current_score:
                        classes[class_a].remove(s_obj)
                        classes[class_b].remove(candidate)
                        classes[class_b].append(s_obj)
                        classes[class_a].append(candidate)
                        current_score = new_score
                        improved = True
                        break
                    else:
                        student_to_class[sid] = class_a  # rückgängig
                        student_to_class[cid] = class_b

                if improved:
                    break
            if improved:
                break

        if not improved:
            # --- Verletzte Getrennt-Wünsche ---
            for sid, wlist in list(wish_dict.items()):
                for w in wlist:
                    if w['wish_type'] != 'separated':
                        continue
                    rid = w['related_student_id']
                    if not rid:
                        continue
                    if sid not in student_to_class or rid not in student_to_class:
                        continue
                    if student_to_class[sid] != student_to_class[rid]:
                        continue

                    s_obj = all_students.get(sid)
                    if not s_obj:
                        continue
                    if sid in pinned_student_ids:
                        continue
                    class_a = student_to_class[sid]

                    for target in range(num_classes):
                        if target == class_a:
                            continue
                        if len(classes[target]) >= MAX_CLASS_SIZE:
                            continue
                        if not ib_move_allowed(s_obj, classes[class_a], classes[target]):
                            continue
                        if not sport_move_allowed(s_obj, class_a, target):
                            continue
                        student_to_class[sid] = target
                        new_score = total_wish_score()
                        if new_score > current_score:
                            classes[class_a].remove(s_obj)
                            classes[target].append(s_obj)
                            current_score = new_score
                            improved = True
                            break
                        else:
                            student_to_class[sid] = class_a

                    if improved:
                        break
                if improved:
                    break

        if not improved:
            break  # Keine weitere Verbesserung möglich

    return current_score


def generate_class_assignment(students, wishes, num_classes, seed, options, base_assignment=None):
    """Intelligente Klasseneinteilung generieren"""
    random.seed(seed)

    # Schüler in Liste konvertieren
    student_list = [dict(s) for s in students]
    random.shuffle(student_list)

    # Spezialklassen: Mapping von Klassen-Indizes zu Spezialtypen
    specialized_mapping = {}
    specialized_classes = options.get('specialized_classes', {})
    class_idx = 0

    # Sport-Klassen zuerst
    sport_count = specialized_classes.get('sport', 0)
    for i in range(sport_count):
        if class_idx < num_classes:
            specialized_mapping[class_idx] = 'sport'
            class_idx += 1

    # Custom-Klassen danach
    custom_count = specialized_classes.get('custom', 0)
    custom_name = specialized_classes.get('custom_name', '')
    for i in range(custom_count):
        if class_idx < num_classes:
            specialized_mapping[class_idx] = 'custom'
            class_idx += 1

    # Bei Sportklassen: Schüler mit Sport-Interesse zuerst sortieren
    if sport_count > 0:
        interested = [s for s in student_list if s.get('sport_interesse')]
        not_interested = [s for s in student_list if not s.get('sport_interesse')]
        student_list = interested + not_interested

    # Bei Sportklasse: sportliche Schüler zuerst verteilen (backward compatibility)
    if options.get('sportklasse', False):
        sportliche = [s for s in student_list if s.get('sportlich')]
        andere = [s for s in student_list if not s.get('sportlich')]
        student_list = sportliche + andere

    # Klassen initialisieren
    classes = [[] for _ in range(num_classes)]

    # Geschlechterzähler pro Klasse
    gender_count = [{'m': 0, 'w': 0} for _ in range(num_classes)]

    # Wohnort-Zähler pro Klasse (für Schulweg-Gruppierung)
    wohnort_count = [{} for _ in range(num_classes)]

    # Städte-Zähler pro Klasse (für Städte-basierte Gruppierung)
    city_count = [{} for _ in range(num_classes)]

    # PLZ-Zähler pro Klasse (für PLZ-basierte Gruppierung)
    plz_count = [{} for _ in range(num_classes)]

    # Schulform-Zähler pro Klasse
    schulform_count = [{'H': 0, 'R': 0, 'G': 0, 'IB': 0, '': 0} for _ in range(num_classes)]

    # Religionszähler pro Klasse
    religion_count = [{'ethik': 0, 'katholisch': 0, 'evangelisch': 0, '': 0} for _ in range(num_classes)]

    # Inklusionszähler pro Klasse
    inklusion_count = [0 for _ in range(num_classes)]

    # IB-Zähler pro Klasse
    ib_count = [0 for _ in range(num_classes)]

    # Basis-Einteilung: bestehende Schüler vorplatzieren, nur neue verteilen
    pinned_student_ids = set()
    if base_assignment:
        base_classes = base_assignment.get('classes', [])
        base_student_class = {}
        for cls in base_classes:
            cls_idx = cls['number'] - 1  # number ist 1-basiert → 0-basiert
            for s in cls.get('students', []):
                base_student_class[s['id']] = cls_idx

        remaining = []
        for student in student_list:
            cls_idx = base_student_class.get(student['id'])
            if cls_idx is not None and 0 <= cls_idx < num_classes:
                classes[cls_idx].append(student)
                pinned_student_ids.add(student['id'])
                g = student.get('gender', 'm')
                if g in gender_count[cls_idx]:
                    gender_count[cls_idx][g] += 1
                wohnort = student.get('wohnort', '').strip()
                if wohnort:
                    wohnort_count[cls_idx][wohnort] = wohnort_count[cls_idx].get(wohnort, 0) + 1
                    city = extract_city_from_wohnort(wohnort)
                    if city:
                        city_count[cls_idx][city] = city_count[cls_idx].get(city, 0) + 1
                    plz = extract_plz_from_wohnort(wohnort)
                    if plz:
                        plz_count[cls_idx][plz] = plz_count[cls_idx].get(plz, 0) + 1
                sf = student.get('schulform', '').strip()
                if sf and sf in schulform_count[cls_idx]:
                    schulform_count[cls_idx][sf] += 1
                if sf == 'IB':
                    ib_count[cls_idx] += 1
                rel = student.get('religion', '') or ''
                if rel in religion_count[cls_idx]:
                    religion_count[cls_idx][rel] += 1
                if student.get('special_needs', ''):
                    inklusion_count[cls_idx] += 1
            else:
                remaining.append(student)
        student_list = remaining

    # Wünsche als Dictionary organisieren
    # Sportklasse hat Priorität: Wünsche zwischen Sport- und Nicht-Sport-Schülern ignorieren,
    # da diese sowieso in verschiedene Klassen kommen.
    sport_student_ids = set()
    if sport_count > 0:
        sport_student_ids = {s['id'] for s in students if s['sport_interesse']}

    wish_dict = {}
    for wish in wishes:
        student_id = wish['student_id']
        related_id = wish['related_student_id']
        if sport_student_ids and related_id:
            s_is_sport = student_id in sport_student_ids
            r_is_sport = related_id in sport_student_ids
            if s_is_sport != r_is_sport:
                continue  # Wunsch ignorieren: einer in Sportklasse, der andere nicht
        if student_id not in wish_dict:
            wish_dict[student_id] = []
        wish_dict[student_id].append(wish)

    # Rückwärts-Wunsch-Dict: related_student_id → Wünsche anderer Schüler über diesen Schüler
    # Ermöglicht bidirektionale Wunsch-Prüfung (wichtig wenn Wunschsteller zuerst platziert wurde)
    reverse_wish_dict = {}
    for sid, student_wishes in wish_dict.items():
        for wish in student_wishes:
            related_id = wish['related_student_id']
            if related_id:
                if related_id not in reverse_wish_dict:
                    reverse_wish_dict[related_id] = []
                reverse_wish_dict[related_id].append(wish)

    # Smarte Anfangs-Reihenfolge: Schüler die von vielen anderen gewünscht werden zuerst platzieren.
    # Wenn sie in einer Klasse sind, finden ihre Freunde sie durch den bidirektionalen Check.
    wish_target_count = {}
    for wish in wishes:
        rid = wish['related_student_id']
        if rid:
            wish_target_count[rid] = wish_target_count.get(rid, 0) + 1

    if not (sport_count > 0 or options.get('sportklasse', False)):
        # Sortiere nach Anzahl der Wünsche über diesen Schüler (absteigende), innerhalb gleicher Gruppe zufällig
        student_list.sort(key=lambda s: -wish_target_count.get(s['id'], 0))
        # Innerhalb gleicher Gruppen zufällig mischen (für Varietät zwischen Vorschlägen)
        from itertools import groupby
        shuffled = []
        for _, grp in groupby(student_list, key=lambda s: wish_target_count.get(s['id'], 0)):
            grp_list = list(grp)
            random.shuffle(grp_list)
            shuffled.extend(grp_list)
        student_list = shuffled

    # IB-Vorverteilung: Schüler deterministisch in Gruppen vorverteilen (garantiert kein Einzelkämpfer)
    if options.get('ib_min', 0) > 0 and options.get('ib_max', 0) > 0:
        ib_min = options.get('ib_min', 0)
        ib_max = options.get('ib_max', 0)
        ib_students_pre = [s for s in student_list if s.get('schulform') == 'IB']
        student_list = [s for s in student_list if s.get('schulform') != 'IB']
        random.shuffle(ib_students_pre)

        num_ib = len(ib_students_pre)
        num_ib_classes = min(num_classes, num_ib // ib_min) if ib_min > 0 else num_classes

        sport_class_indices = {ci for ci, t in specialized_mapping.items() if t == 'sport'}

        if num_ib_classes > 0:
            for idx, student in enumerate(ib_students_pre):
                target_class = idx % num_ib_classes
                # IB ohne Sportklassen-Hacken nicht in Sportklassen einteilen
                if not student.get('sport_interesse') and target_class in sport_class_indices:
                    for alt in range(num_ib_classes):
                        if alt not in sport_class_indices and (ib_max == 0 or ib_count[alt] < ib_max):
                            target_class = alt
                            break
                # Falls Maximum erreicht, nächste Klasse mit Platz suchen
                if ib_count[target_class] >= ib_max:
                    for alt in range(num_ib_classes):
                        if ib_count[alt] < ib_max:
                            target_class = alt
                            break
                classes[target_class].append(student)
                ib_count[target_class] += 1
                gender = student.get('gender', 'm')
                if gender in gender_count[target_class]:
                    gender_count[target_class][gender] += 1
                wohnort = student.get('wohnort', '').strip()
                if wohnort:
                    wohnort_count[target_class][wohnort] = wohnort_count[target_class].get(wohnort, 0) + 1
                    city = extract_city_from_wohnort(wohnort)
                    if city:
                        city_count[target_class][city] = city_count[target_class].get(city, 0) + 1
                    plz = extract_plz_from_wohnort(wohnort)
                    if plz:
                        plz_count[target_class][plz] = plz_count[target_class].get(plz, 0) + 1
                schulform = student.get('schulform', '').strip()
                if schulform and schulform in schulform_count[target_class]:
                    schulform_count[target_class][schulform] += 1
                religion = student.get('religion', '') or ''
                if religion in religion_count[target_class]:
                    religion_count[target_class][religion] += 1
                if student.get('special_needs', ''):
                    inklusion_count[target_class] += 1
        else:
            # Zu wenige IB-Schüler für ib_min → normal verteilen
            student_list = ib_students_pre + student_list

    # Schüler auf Klassen verteilen
    for student in student_list:
        best_class = find_best_class(student, classes, gender_count, wohnort_count, city_count, plz_count, schulform_count, religion_count, inklusion_count, ib_count, wish_dict, num_classes, options, specialized_mapping, reverse_wish_dict)
        classes[best_class].append(student)

        # Geschlechterzähler aktualisieren
        gender = student.get('gender', 'm')
        if gender in gender_count[best_class]:
            gender_count[best_class][gender] += 1

        # Wohnort-Zähler aktualisieren
        wohnort = student.get('wohnort', '').strip()
        if wohnort:
            if wohnort in wohnort_count[best_class]:
                wohnort_count[best_class][wohnort] += 1
            else:
                wohnort_count[best_class][wohnort] = 1

            # Städte-Zähler aktualisieren (PLZ + Stadt)
            city = extract_city_from_wohnort(wohnort)
            if city:
                if city in city_count[best_class]:
                    city_count[best_class][city] += 1
                else:
                    city_count[best_class][city] = 1

            # PLZ-Zähler aktualisieren (nur PLZ)
            plz = extract_plz_from_wohnort(wohnort)
            if plz:
                if plz in plz_count[best_class]:
                    plz_count[best_class][plz] += 1
                else:
                    plz_count[best_class][plz] = 1

        # Schulform-Zähler aktualisieren
        schulform = student.get('schulform', '').strip()
        if schulform:
            if schulform in schulform_count[best_class]:
                schulform_count[best_class][schulform] += 1

        # Religionszähler aktualisieren
        religion = student.get('religion', '') or ''
        if religion in religion_count[best_class]:
            religion_count[best_class][religion] += 1

        # Inklusionszähler aktualisieren
        if student.get('special_needs', ''):
            inklusion_count[best_class] += 1

        # IB-Zähler aktualisieren
        if student.get('schulform') == 'IB':
            ib_count[best_class] += 1

    # Post-Processing: Wunsch-Erfüllungsrate durch iterativen Tausch verbessern
    optimize_assignment_wishes(classes, wish_dict, options=options, pinned_student_ids=pinned_student_ids, specialized_mapping=specialized_mapping)

    # Statistiken nach Optimierung neu berechnen (Klassen-Zusammensetzung hat sich geändert)
    gender_count = [{'m': 0, 'w': 0} for _ in range(num_classes)]
    wohnort_count = [{} for _ in range(num_classes)]
    schulform_count = [{'H': 0, 'R': 0, 'G': 0, 'IB': 0, '': 0} for _ in range(num_classes)]
    religion_count = [{'ethik': 0, 'katholisch': 0, 'evangelisch': 0, '': 0} for _ in range(num_classes)]
    inklusion_count = [0] * num_classes
    ib_count = [0] * num_classes

    for i, cls in enumerate(classes):
        for s in cls:
            g = s.get('gender', 'm')
            if g in gender_count[i]:
                gender_count[i][g] += 1
            wo = s.get('wohnort', '').strip()
            if wo:
                wohnort_count[i][wo] = wohnort_count[i].get(wo, 0) + 1
            sf = s.get('schulform', '').strip()
            if sf in schulform_count[i]:
                schulform_count[i][sf] += 1
            rel = s.get('religion', '') or ''
            if rel in religion_count[i]:
                religion_count[i][rel] += 1
            if s.get('special_needs', ''):
                inklusion_count[i] += 1
            if s.get('schulform') == 'IB':
                ib_count[i] += 1

    # Ergebnis formatieren
    result = {
        'classes': [],
        'statistics': {
            'total_students': len(student_list),
            'num_classes': num_classes,
            'gender_distribution': gender_count
        }
    }

    for i, cls in enumerate(classes):
        sport_count = sum(1 for s in cls if s.get('sportlich'))

        # Wohnorte vereinfachen: PLZ + Stadt extrahieren und gruppieren
        city_count = {}
        for wohnort, count in wohnort_count[i].items():
            # Extrahiere PLZ (erste 5 Ziffern) und Stadt
            import re
            match = re.search(r'(\d{5})\s+([^,]+)', wohnort)
            if match:
                plz = match.group(1)
                stadt = match.group(2).strip()
                # Entferne Klammern wie "(Taunus)"
                stadt = re.sub(r'\s*\([^)]*\)', '', stadt).strip()
                city_key = f"{plz} {stadt}"
                if city_key in city_count:
                    city_count[city_key] += count
                else:
                    city_count[city_key] = count

        # Bestimme Spezialklassen-Typ
        special_type = specialized_mapping.get(i, None)
        is_sportklasse = options.get('sportklasse', False) and i == 0  # Backward compatibility
        custom_name = options.get('specialized_classes', {}).get('custom_name', '') if special_type == 'custom' else ''

        result['classes'].append({
            'number': i + 1,
            'students': cls,
            'count': len(cls),
            'gender_count': gender_count[i],
            'wohnort_count': wohnort_count[i],
            'city_count': city_count,
            'schulform_count': schulform_count[i],
            'religion_count': religion_count[i],
            'inklusion_count': inklusion_count[i],
            'ib_count': ib_count[i],
            'sport_count': sport_count,
            'is_sportklasse': is_sportklasse,
            'special_type': special_type,
            'custom_name': custom_name
        })

    return result

def find_best_class(student, classes, gender_count, wohnort_count, city_count, plz_count, schulform_count, religion_count, inklusion_count, ib_count, wish_dict, num_classes, options, specialized_mapping=None, reverse_wish_dict=None):
    """Beste Klasse für einen Schüler finden"""
    scores = []
    ib_min = options.get('ib_min', 0)
    ib_max = options.get('ib_max', 0)
    is_ib_student = student.get('schulform') == 'IB'
    if specialized_mapping is None:
        specialized_mapping = {}
    if reverse_wish_dict is None:
        reverse_wish_dict = {}

    ib_class_size = options.get('ib_class_size', 0)
    student_special = student.get('special_needs', '')
    is_ikl = bool(student.get('ikl', 0))

    for i, cls in enumerate(classes):
        score = 0

        # HARTE GRENZE: Klassengröße
        # IB-Klassen bekommen ein reduziertes Maximum (konfigurierbar)
        effective_max = MAX_CLASS_SIZE
        if ib_class_size > 0 and ib_count[i] > 0:
            effective_max = ib_class_size
        if len(cls) >= effective_max:
            score -= 10000
            scores.append(score)
            continue

        # Größenausgleich (bevorzuge kleinere Klassen)
        avg_size = sum(len(c) for c in classes) / num_classes
        size_diff = len(cls) - avg_size
        score -= size_diff * 10

        # Spezialklassen: Interesse-basiertes Scoring
        if i in specialized_mapping:
            special_type = specialized_mapping[i]
            # Nur für Sport-Klassen (Custom hat kein Interesse-Feld)
            if special_type == 'sport':
                if student.get('sport_interesse'):
                    score += 500  # Sehr starker Bonus: Sport-Schüler gehören in Sportklasse
                else:
                    score -= 5000  # Harte Sperre: kein Sportklassen-Hacken → nie in Sportklasse

        # Wenn Schüler Sport-Interesse hat aber nicht in Sport-Spezialklasse
        if student.get('sport_interesse'):
            if specialized_mapping.get(i) != 'sport':
                score -= 200  # Starke Strafe: Sport-Schüler sollen nicht in Normal-Klassen

        # Sportklasse (backward compatibility): Klasse 1 (Index 0) für sportliche Schüler
        if options.get('sportklasse', False):
            if student.get('sportlich'):
                if i == 0:
                    score += 50  # Sportliche stark in Klasse 1
                else:
                    score -= 30  # Sportliche weg von anderen Klassen
            else:
                if i == 0:
                    score -= 30  # Nicht-sportliche weg von Klasse 1

        # IB Max-Grenze einhalten (Vorverteilung deckt den Normalfall ab)
        if ib_max > 0 and is_ib_student and ib_count[i] >= ib_max:
            score -= 1000  # Harte Blockade - Maximum erreicht

        # Förderbedarf Sprache: möglichst in einer Klasse bündeln
        # ESE (sozial_emotional) wird NICHT gebündelt
        if student_special == 'sprache':
            sprache_in_cls = sum(1 for s in cls if s.get('special_needs') == 'sprache')
            score += sprache_in_cls * 80  # Starker Bonus je weiteren Sprache-Schüler in Klasse

        # IKL: möglichst NICHT in Klassen mit IB-Schülern
        if is_ikl:
            if ib_count[i] > 0:
                score -= 500  # Starke Strafe für IKL in IB-Klasse
            # IKL-Schüler auf Klassen verteilen (nicht alle in eine Klasse)
            ikl_in_cls = sum(1 for s in cls if s.get('ikl', 0))
            score -= ikl_in_cls * 60  # Penalty für Konzentration von IKL in einer Klasse

        # Geschlechterbalance (SEHR WICHTIG - höchste Priorität)
        if options.get('gender_balance', True):
            gender = student.get('gender', 'm')
            if gender in gender_count[i]:
                total = sum(gender_count[i].values())
                if total > 0:
                    gender_ratio = gender_count[i][gender] / total
                    score -= gender_ratio * 15  # Erhöht von 5 auf 15 für höhere Priorität

        # Schulweg-Gruppierung (WICHTIG - Schüler aus gleicher Stadt/PLZ zusammen)
        if options.get('schulweg_gruppe', True):
            wohnort = student.get('wohnort', '').strip()
            if wohnort:
                # Stadt-Gruppierung (PLZ + Stadtname)
                city = extract_city_from_wohnort(wohnort)
                if city and city in city_count[i]:
                    # HOHER Bonus für gleiche Stadt (Fahrgemeinschaften, soziale Verbindungen)
                    score += city_count[i][city] * 20

                # PLZ-Gruppierung (zusätzlicher Bonus wenn gleiche PLZ aber andere Stadt)
                plz = extract_plz_from_wohnort(wohnort)
                if plz and plz in plz_count[i]:
                    # Mittlerer Bonus für gleiche PLZ (regionale Nähe)
                    score += plz_count[i][plz] * 10

        # Schulform-Verteilung
        if options.get('schulform_balance', True):
            schulform = student.get('schulform', '').strip()
            if schulform:
                total = sum(schulform_count[i].values())
                if total > 0:
                    schulform_ratio = schulform_count[i].get(schulform, 0) / total
                    score -= schulform_ratio * 8

        # Religion verteilen (ZWEITRANGIG - reduzierte Priorität)
        if options.get('religion_distribute', False):
            religion = student.get('religion', '') or ''
            if religion and religion in religion_count[i]:
                total = sum(religion_count[i].values())
                if total > 0:
                    rel_ratio = religion_count[i][religion] / total
                    score -= rel_ratio * 2  # Reduziert von 5 auf 2

        # Religion gruppieren (ZWEITRANGIG - reduzierte Priorität)
        if options.get('religion_group', False):
            religion = student.get('religion', '') or ''
            if religion and religion in religion_count[i]:
                if religion_count[i][religion] > 0:
                    score += 2  # Reduziert von 5 auf 2

        # Religion-Bündelung (Ethik mit Konfessionen)
        if options.get('religion_bundle', False):
            religion = student.get('religion', '') or ''
            if religion == 'ethik':
                konfession_count = religion_count[i].get('katholisch', 0) + religion_count[i].get('evangelisch', 0)
                if konfession_count > 0:
                    score += 15  # Bonus für Vielfalt
                else:
                    score -= 20  # Strafe für Isolation (reine Ethik-Klasse vermeiden)
            elif religion in ['katholisch', 'evangelisch']:
                if religion_count[i].get('ethik', 0) > 0:
                    score += 8  # Bonus wenn Ethik vorhanden

        # Elternwünsche berücksichtigen (SEHR WICHTIG – höchste Priorität nach Klassengröße)
        if options.get('parent_wishes', True):
            student_id = student['id']

            # Eigene Wünsche des Schülers prüfen
            for wish in wish_dict.get(student_id, []):
                wish_type = wish['wish_type']
                related_id = wish['related_student_id']
                if not related_id:
                    continue
                related_in_class = any(s['id'] == related_id for s in cls)
                if wish_type == 'together' and related_in_class:
                    score += 150   # Sehr hoher Bonus: "mit Freund/in zusammen"
                elif wish_type == 'separated' and related_in_class:
                    score -= 5000  # Harte Sperre: "Auf keinen Fall mit" (quasi unüberwindbar)

            # Rückwärts-Wünsche: Schüler X hat Wunsch über diesen Schüler
            # (wichtig wenn X bereits platziert wurde, bevor dieser Schüler drankommt)
            for wish in reverse_wish_dict.get(student_id, []):
                wish_type = wish['wish_type']
                wisher_id = wish['student_id']
                wisher_in_class = any(s['id'] == wisher_id for s in cls)
                if wish_type == 'together' and wisher_in_class:
                    score += 150   # Wunschsteller ist bereits in dieser Klasse → Bonus
                elif wish_type == 'separated' and wisher_in_class:
                    score -= 500   # Wunschsteller ist bereits in dieser Klasse → Strafe

        scores.append(score)

    # Absolute Sperre: Nicht-Sport-Schüler dürfen NIEMALS in Sportklasse
    # (gilt auch wenn Sportklasse durch "voll"-Check schon -10000 hat und die Strafe in der
    #  Hauptschleife durch `continue` übersprungen wurde)
    if not student.get('sport_interesse') and specialized_mapping:
        for sci, st in specialized_mapping.items():
            if st == 'sport' and sci < len(scores):
                scores[sci] -= 100000

    # Klasse mit höchstem Score auswählen
    best_class = scores.index(max(scores))
    return best_class

# ─────────────────────────────────────────────────────────────
# Update-System
# ─────────────────────────────────────────────────────────────

def _version_tuple(v):
    try:
        return tuple(int(x) for x in str(v).strip().split('.'))
    except Exception:
        return (0, 0, 0)

def _get_update_backup_dir():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'update_backup')

def _do_rollback():
    """Stellt die letzte gesicherte Version wieder her."""
    import shutil
    app_dir = os.path.dirname(os.path.abspath(__file__))
    backup_dir = _get_update_backup_dir()
    if not os.path.exists(backup_dir):
        raise FileNotFoundError('Kein Backup vorhanden.')
    for item in os.listdir(backup_dir):
        src = os.path.join(backup_dir, item)
        dst = os.path.join(app_dir, item)
        if os.path.isdir(src):
            if os.path.exists(dst):
                shutil.rmtree(dst)
            shutil.copytree(src, dst)
        else:
            shutil.copy2(src, dst)

def _fetch_github_text(url, timeout=15):
    import urllib.request
    req = urllib.request.Request(url, headers={'User-Agent': 'KlasseneinteilungApp-Updater/1.0'})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode('utf-8', errors='replace')

def _parse_new_changelog(changelog_text, current_ver):
    """Gibt nur die Abschnitte zurück, die neuer als current_ver sind."""
    sections = []
    current_section = []
    in_new_section = False
    for line in changelog_text.splitlines():
        if line.startswith('## Version '):
            if current_section and in_new_section:
                sections.append('\n'.join(current_section))
            current_section = [line]
            m = re.search(r'## Version (\S+)', line)
            if m:
                in_new_section = _version_tuple(m.group(1)) > _version_tuple(current_ver)
            else:
                in_new_section = False
        else:
            if current_section:
                current_section.append(line)
    if current_section and in_new_section:
        sections.append('\n'.join(current_section))
    return sections


@app.route('/admin/update')
@admin_required
def check_update():
    """Update-Seite: prüft ob eine neuere Version auf GitHub verfügbar ist (via Releases API)."""
    import json as _json, urllib.request
    github_version = None
    error = None

    try:
        api_url = f'https://api.github.com/repos/{GITHUB_REPO}/releases/latest'
        req = urllib.request.Request(api_url, headers={
            'User-Agent': 'KlasseneinteilungApp-Updater/1.0',
            'Accept': 'application/vnd.github.v3+json',
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode('utf-8'))
        tag = data.get('tag_name', '')
        m = re.match(r'^v?(\d+\.\d+\.\d+)$', tag)
        if m:
            github_version = m.group(1)
        else:
            error = 'Konnte Version nicht aus GitHub lesen.'
    except Exception as e:
        error = f'GitHub nicht erreichbar: {e}'

    update_available = (github_version is not None and
                        _version_tuple(github_version) > _version_tuple(__version__))

    has_backup = os.path.exists(_get_update_backup_dir())

    return render_template('update.html',
                           current_version=__version__,
                           github_version=github_version,
                           update_available=update_available,
                           error=error,
                           has_backup=has_backup)


@app.route('/admin/update/apply', methods=['POST'])
@admin_required
@limiter.limit("5 per hour")
def apply_update():
    """Lädt die neueste Version von GitHub herunter und installiert sie."""
    import urllib.request, zipfile, shutil

    app_dir = os.path.dirname(os.path.abspath(__file__))
    backup_dir = _get_update_backup_dir()

    # Schutzliste: diese Einträge werden beim Update nie überschrieben
    PRESERVE_FILES = {'.env', '.initial_password', 'klasseneinteilung.db'}
    PRESERVE_DIRS  = {'flask_session', 'python-portable', 'update_backup', '__pycache__', 'user_data', 'venv'}

    # 1. Backup anlegen
    try:
        if os.path.exists(backup_dir):
            shutil.rmtree(backup_dir)
        os.makedirs(backup_dir, exist_ok=True)
        for item in ['app.py', 'requirements.txt', 'VERSION_HISTORY.md']:
            src = os.path.join(app_dir, item)
            if os.path.exists(src):
                shutil.copy2(src, os.path.join(backup_dir, item))
        for folder in ['templates', 'static']:
            src = os.path.join(app_dir, folder)
            if os.path.exists(src):
                shutil.copytree(src, os.path.join(backup_dir, folder))
    except Exception as e:
        flash(f'Backup fehlgeschlagen — Update abgebrochen: {e}', 'danger')
        return redirect(url_for('check_update'))

    # 2. ZIP von GitHub laden
    try:
        req = urllib.request.Request(
            GITHUB_ZIP_URL,
            headers={'User-Agent': 'KlasseneinteilungApp-Updater/1.0'}
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            zip_data = resp.read()
    except Exception as e:
        flash(f'Download fehlgeschlagen: {e}', 'danger')
        return redirect(url_for('check_update'))

    # 3. ZIP extrahieren und Dateien ersetzen
    try:
        with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
            names = zf.namelist()
            if not names:
                raise ValueError('ZIP ist leer.')
            # GitHub ZIPs haben ein Top-Level-Verzeichnis: "repo-main/"
            prefix = names[0].split('/')[0] + '/'

            for member in names:
                if not member.startswith(prefix):
                    continue
                rel = member[len(prefix):]  # Relativer Pfad im Zielverzeichnis
                if not rel:
                    continue
                top = rel.split('/')[0]
                if top in PRESERVE_FILES or top in PRESERVE_DIRS:
                    continue
                dest = os.path.join(app_dir, rel)
                # Zip-Slip-Schutz: dest muss innerhalb von app_dir liegen
                real_dest = os.path.realpath(dest)
                if real_dest != os.path.realpath(app_dir) and \
                   not real_dest.startswith(os.path.realpath(app_dir) + os.sep):
                    continue
                # Geschützte Pfade auch bei verschachtelten Traversal-Versuchen wahren
                rel_top = os.path.relpath(real_dest, os.path.realpath(app_dir)).split(os.sep)[0]
                if rel_top in PRESERVE_FILES or rel_top in PRESERVE_DIRS:
                    continue
                if member.endswith('/'):
                    os.makedirs(dest, exist_ok=True)
                else:
                    os.makedirs(os.path.dirname(dest), exist_ok=True)
                    with zf.open(member) as src_f, open(dest, 'wb') as dst_f:
                        dst_f.write(src_f.read())

        # Neustart: SIGHUP mit 2s Delay damit HTTP-Antwort zuerst gesendet wird
        try:
            import signal as _signal, threading
            master_pid = os.getppid()
            def _delayed_sighup():
                import time; time.sleep(2)
                try: os.kill(master_pid, _signal.SIGHUP)
                except Exception: pass
            threading.Thread(target=_delayed_sighup, daemon=True).start()
            flash('✅ Update erfolgreich installiert! Die App startet in wenigen Sekunden automatisch neu.', 'success')
        except Exception:
            flash('✅ Update erfolgreich installiert! Bitte starte die App neu (Fenster schließen → PORTABLE-START.bat).', 'success')
    except Exception as e:
        # Automatischer Rollback
        try:
            _do_rollback()
            flash(f'❌ Update fehlgeschlagen ({e}). Rollback erfolgreich — vorherige Version wiederhergestellt. Bitte App neu starten.', 'warning')
        except Exception as rb_err:
            flash(f'❌ Update UND Rollback fehlgeschlagen: {e} / {rb_err}. Manuelle Wiederherstellung erforderlich.', 'danger')

    return redirect(url_for('check_update'))


@app.route('/admin/update/rollback', methods=['POST'])
@admin_required
@limiter.limit("5 per hour")
def rollback_update():
    """Stellt die gesicherte Version manuell wieder her."""
    try:
        _do_rollback()
        flash('↩️ Rollback erfolgreich. Bitte starte die App neu (Fenster schließen → PORTABLE-START.bat).', 'success')
    except Exception as e:
        flash(f'Rollback fehlgeschlagen: {e}', 'danger')
    return redirect(url_for('check_update'))


@app.route('/wizard')
@login_required
def wizard():
    """Wizard starten oder fortsetzen"""
    # Wizard aktivieren
    session['wizard_active'] = True
    if 'wizard_step' not in session:
        session['wizard_step'] = 1

    return redirect(url_for('wizard_step', step=session['wizard_step']))

@app.route('/wizard/<int:step>')
@login_required
def wizard_step(step):
    """Wizard-Schritt anzeigen"""
    if not session.get('wizard_active'):
        return redirect(url_for('dashboard'))

    session['wizard_step'] = step

    db = get_db()
    cursor = db.cursor()

    # Statistiken für alle Schritte
    cursor.execute('SELECT COUNT(*) as count FROM students')
    student_count = cursor.fetchone()['count']

    cursor.execute('SELECT COUNT(*) as count FROM parent_wishes')
    wish_count = cursor.fetchone()['count']

    db.close()

    return render_template('wizard.html',
                         step=step,
                         student_count=student_count,
                         wish_count=wish_count)

@app.route('/wizard/cancel')
@login_required
def wizard_cancel():
    """Wizard abbrechen"""
    session.pop('wizard_active', None)
    session.pop('wizard_step', None)
    flash('Wizard abgebrochen.', 'info')
    return redirect(url_for('dashboard'))

@app.route('/wizard/restart')
@login_required
def wizard_restart():
    """Wizard neu starten"""
    session['wizard_active'] = True
    session['wizard_step'] = 1
    flash('Wizard wurde neu gestartet.', 'info')
    return redirect(url_for('wizard_step', step=1))

@app.route('/wizard/complete')
@login_required
def wizard_complete():
    """Wizard abschließen"""
    session.pop('wizard_active', None)
    session.pop('wizard_step', None)
    flash('Glückwunsch! Sie haben erfolgreich eine Klasseneinteilung erstellt! 🎉', 'success')
    return redirect(url_for('assignments'))

@app.route('/assignments')
@login_required
def assignments():
    """Gespeicherte Einteilungen anzeigen"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM class_assignments ORDER BY created_at DESC')
    assignments = cursor.fetchall()
    db.close()

    return render_template('assignments.html', assignments=assignments)

@app.route('/assignments/compare')
@login_required
def compare_assignments():
    """Zwei gespeicherte Einteilungen vergleichen"""
    import json as _json

    a_id = request.args.get('a', type=int)
    b_id = request.args.get('b', type=int)

    if not a_id or not b_id or a_id == b_id:
        flash('Bitte genau zwei verschiedene Einteilungen auswählen.', 'warning')
        return redirect(url_for('assignments'))

    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM class_assignments WHERE id = ?', (a_id,))
    assignment_a = cursor.fetchone()
    cursor.execute('SELECT * FROM class_assignments WHERE id = ?', (b_id,))
    assignment_b = cursor.fetchone()
    db.close()

    if not assignment_a or not assignment_b:
        flash('Einteilung nicht gefunden.', 'danger')
        return redirect(url_for('assignments'))

    def build_map(proposal):
        m = {}
        for cls in proposal.get('classes', []):
            for s in cls.get('students', []):
                m[s['id']] = {'class_name': cls['name'], 'firstname': s.get('firstname', ''), 'lastname': s.get('lastname', '')}
        return m

    map_a = build_map(_json.loads(assignment_a['data']))
    map_b = build_map(_json.loads(assignment_b['data']))

    all_ids = set(map_a.keys()) | set(map_b.keys())
    rows = []
    for sid in all_ids:
        info = map_a.get(sid) or map_b.get(sid)
        a_class = map_a[sid]['class_name'] if sid in map_a else '—'
        b_class = map_b[sid]['class_name'] if sid in map_b else '—'
        rows.append({
            'name': f"{info['lastname']}, {info['firstname']}",
            'a_class': a_class,
            'b_class': b_class,
            'changed': a_class != b_class,
            'only_in_a': sid not in map_b,
            'only_in_b': sid not in map_a,
        })
    rows.sort(key=lambda r: r['name'])
    changed_count = sum(1 for r in rows if r['changed'])

    return render_template('compare_assignments.html',
                           assignment_a=assignment_a,
                           assignment_b=assignment_b,
                           rows=rows,
                           changed_count=changed_count,
                           total_count=len(rows))

@app.route('/assignments/<int:assignment_id>/delete', methods=['POST'])
@login_required
def delete_assignment(assignment_id):
    """Gespeicherte Einteilung löschen"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT id FROM class_assignments WHERE id = ?', (assignment_id,))
    if not cursor.fetchone():
        db.close()
        flash('Einteilung nicht gefunden.', 'danger')
        return redirect(url_for('assignments'))
    cursor.execute('DELETE FROM class_assignments WHERE id = ?', (assignment_id,))
    db.commit()
    db.close()
    flash('Einteilung wurde gelöscht.', 'success')
    return redirect(url_for('assignments'))

@app.route('/assignments/<int:assignment_id>')
@login_required
def view_assignment(assignment_id):
    """Einzelne gespeicherte Einteilung anzeigen"""
    import json

    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM class_assignments WHERE id = ?', (assignment_id,))
    assignment = cursor.fetchone()
    db.close()

    if not assignment:
        flash('Einteilung nicht gefunden.', 'danger')
        return redirect(url_for('assignments'))

    # Daten aus JSON parsen
    proposal = json.loads(assignment['data'])

    # Als einzelnes Proposal anzeigen (wie in generate.html, aber nur eines)
    return render_template('view_assignment.html',
                         assignment=assignment,
                         proposal=proposal,
                         num_classes=len(proposal['classes']))

@app.route('/assignments/<int:assignment_id>/update', methods=['POST'])
@login_required
def update_assignment(assignment_id):
    """Gespeicherte Einteilung nach Drag-&-Drop-Änderungen aktualisieren"""
    import json
    arrangement_json = request.form.get('arrangement', '')
    if not arrangement_json:
        flash('Keine Änderungen übermittelt.', 'warning')
        return redirect(url_for('view_assignment', assignment_id=assignment_id))

    try:
        arrangement = json.loads(arrangement_json)  # {student_id: class_number_str}
    except (ValueError, TypeError):
        flash('Ungültige Daten.', 'danger')
        return redirect(url_for('view_assignment', assignment_id=assignment_id))

    db = get_db()
    row = db.execute('SELECT data FROM class_assignments WHERE id = ?', (assignment_id,)).fetchone()
    if not row:
        db.close()
        flash('Einteilung nicht gefunden.', 'danger')
        return redirect(url_for('assignments'))

    proposal = json.loads(row['data'])

    # Alle Schüler aus der Einteilung als Lookup aufbauen
    all_students = {}
    for cls in proposal['classes']:
        for s in cls['students']:
            all_students[str(s['id'])] = s

    # Klassen neu befüllen basierend auf der übermittelten Zuordnung
    for cls in proposal['classes']:
        cls_num = str(cls['number'])
        new_students = [all_students[sid] for sid, cnum in arrangement.items()
                        if cnum == cls_num and sid in all_students]
        cls['students'] = new_students
        cls['count'] = len(new_students)
        cls['gender_count'] = {
            'm': sum(1 for s in new_students if s.get('gender') == 'm'),
            'w': sum(1 for s in new_students if s.get('gender') == 'w')
        }
        cls['schulform_count'] = {'H': 0, 'R': 0, 'G': 0, 'IB': 0, '': 0}
        for s in new_students:
            sf = s.get('schulform', '') or ''
            if sf in cls['schulform_count']:
                cls['schulform_count'][sf] += 1
        cls['ib_count'] = sum(1 for s in new_students if s.get('schulform') == 'IB')
        cls['sport_count'] = sum(1 for s in new_students if s.get('sportlich'))
        cls['inklusion_count'] = sum(1 for s in new_students if s.get('special_needs'))

    db.execute('UPDATE class_assignments SET data = ? WHERE id = ?',
               (json.dumps(proposal), assignment_id))
    db.commit()
    db.close()

    flash('Einteilung erfolgreich gespeichert.', 'success')
    return redirect(url_for('view_assignment', assignment_id=assignment_id))


@app.route('/save_assignment', methods=['POST'])
@login_required
def save_assignment():
    """Ausgewählte Einteilung speichern"""
    import json

    proposal_index = int(request.form.get('proposal_index', 0))
    proposals = session.get('last_proposals', [])

    if not proposals or proposal_index >= len(proposals):
        flash('Keine gültige Einteilung gefunden. Bitte generieren Sie zuerst eine Einteilung.', 'danger')
        return redirect(url_for('generate'))

    proposal = proposals[proposal_index]

    # Automatischen Namen generieren
    now = datetime.now()
    assignment_name = f"Einteilung vom {now.strftime('%d.%m.%Y %H:%M')}"

    # In Datenbank speichern
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        INSERT INTO class_assignments (name, data, username, created_at)
        VALUES (?, ?, ?, ?)
    ''', (assignment_name, json.dumps(proposal), session.get('username', ''), now.strftime('%Y-%m-%d %H:%M:%S')))
    db.commit()
    db.close()

    num_classes = len(proposal['classes'])
    num_students = len([s for c in proposal['classes'] for s in c['students']])
    flash(f'Einteilung "{assignment_name}" erfolgreich gespeichert! ({num_classes} Klassen, {num_students} Schüler)', 'success')
    return redirect(url_for('assignments'))

@app.route('/assignments/<int:assignment_id>/export/<format_type>', methods=['POST'])
@login_required
def export_saved_assignment(assignment_id, format_type):
    """Exportiert eine gespeicherte Einteilung"""
    import json

    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT data FROM class_assignments WHERE id = ?', (assignment_id,))
    assignment = cursor.fetchone()
    db.close()

    if not assignment:
        flash('Einteilung nicht gefunden.', 'danger')
        return redirect(url_for('assignments'))

    proposal = json.loads(assignment['data'])

    if format_type == 'excel':
        return generate_excel_export(proposal)
    elif format_type == 'csv':
        return generate_csv_export(proposal)
    elif format_type == 'pdf':
        return generate_pdf_export(proposal)
    else:
        flash('Ungültiges Export-Format.', 'danger')
        return redirect(url_for('view_assignment', assignment_id=assignment_id))

@app.route('/export/<format_type>', methods=['POST'])
@login_required
def export_classes(format_type):
    """Export Klasseneinteilung in verschiedenen Formaten"""
    from flask import make_response, send_file

    try:
        # Hole die Proposals aus der Session
        proposals = session.get('last_proposals', [])
        proposal_index = int(request.form.get('proposal_index', 0))

        if not proposals or proposal_index >= len(proposals):
            flash('Keine gültige Einteilung gefunden. Bitte generieren Sie zuerst eine Einteilung.', 'danger')
            return redirect(url_for('generate'))

        proposal = proposals[proposal_index]

        if format_type == 'excel':
            return generate_excel_export(proposal)
        elif format_type == 'csv':
            return generate_csv_export(proposal)
        elif format_type == 'pdf':
            return generate_pdf_export(proposal)
        else:
            flash('Ungültiges Export-Format.', 'danger')
            return redirect(url_for('generate'))
    except Exception as e:
        flash(f'Fehler beim Export: {str(e)}', 'danger')
        return redirect(url_for('generate'))

def generate_excel_export(proposal):
    """Generiert Excel-Export"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    wb.remove(wb.active)  # Entferne leeres Sheet

    # Erstelle ein Sheet pro Klasse
    for class_data in proposal['classes']:
        class_name = f"5{chr(96 + class_data['number'])}"
        ws = wb.create_sheet(title=class_name)

        # Header
        ws['A1'] = f"Klasse {class_name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')

        # Statistiken
        ws['A3'] = 'Statistiken:'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = f"Gesamt: {class_data['count']} Schüler"
        ws['A5'] = f"Männlich: {class_data['gender_count']['m']}"
        ws['A6'] = f"Weiblich: {class_data['gender_count']['w']}"

        if class_data.get('schulform_count'):
            row = 8
            ws[f'A{row}'] = 'Schulformen:'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for sf, count in class_data['schulform_count'].items():
                if count > 0 and sf:
                    ws[f'A{row}'] = f"{sf}: {count}"
                    row += 1

        # Schülerliste Header
        start_row = 14
        ws[f'A{start_row}'] = 'Nachname'
        ws[f'B{start_row}'] = 'Vorname'
        ws[f'C{start_row}'] = 'Geschlecht'
        ws[f'D{start_row}'] = 'Schulform'
        ws[f'E{start_row}'] = 'IB'
        ws[f'F{start_row}'] = 'Wohnort'

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{start_row}'].font = Font(bold=True)
            ws[f'{col}{start_row}'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Schüler eintragen
        row = start_row + 1
        for student in sorted(class_data['students'], key=lambda s: (s['lastname'], s['firstname'])):
            schulform = student.get('schulform', '')
            is_ib = 'Ja' if schulform == 'IB' else ''
            ws[f'A{row}'] = student['lastname']
            ws[f'B{row}'] = student['firstname']
            ws[f'C{row}'] = student.get('gender', '')
            ws[f'D{row}'] = schulform
            ws[f'E{row}'] = is_ib
            ws[f'F{row}'] = student.get('wohnort', '')
            row += 1

        # Spaltenbreite anpassen
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 30

    # Speichern in BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'klasseneinteilung_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

def generate_csv_export(proposal):
    """Generiert CSV-Export"""
    from flask import make_response
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    writer.writerow(['Klasse', 'Nachname', 'Vorname', 'Geschlecht', 'Schulform', 'IB', 'Wohnort', 'Religion'])

    # Schüler pro Klasse
    for class_data in proposal['classes']:
        class_name = f"5{chr(96 + class_data['number'])}"
        for student in sorted(class_data['students'], key=lambda s: (s['lastname'], s['firstname'])):
            schulform = student.get('schulform', '')
            is_ib = 'Ja' if schulform == 'IB' else ''
            writer.writerow([
                class_name,
                student['lastname'],
                student['firstname'],
                student.get('gender', ''),
                schulform,
                is_ib,
                student.get('wohnort', ''),
                student.get('religion', '')
            ])

    # Response erstellen
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=klasseneinteilung_{datetime.now().strftime("%Y%m%d")}.csv'

    return response

def generate_pdf_export(proposal):
    """Generiert PDF-Export mit ReportLab"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from flask import send_file
    import unicodedata

    def normalize_text(text):
        """Normalisiert Text für PDF-Export (entfernt diakritische Zeichen)"""
        if not text:
            return text
        # NFD = Normalization Form Decomposed - trennt Buchstaben von diakritischen Zeichen
        # Dann filtern wir nur die Basis-Buchstaben heraus
        nfd = unicodedata.normalize('NFD', text)
        return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')

    # PDF erstellen
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    styles = getSampleStyleSheet()

    # Titel-Style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1d1d1f'),
        spaceAfter=20
    )

    # Pro Klasse eine Seite
    for idx, class_data in enumerate(proposal['classes']):
        class_name = f"5{chr(96 + class_data['number'])}"

        # Titel
        elements.append(Paragraph(f"Klasse {class_name}", title_style))

        # Statistiken
        stats_data = [
            ['Gesamt:', f"{class_data['count']} Schueler"],
            ['Maennlich:', str(class_data['gender_count']['m'])],
            ['Weiblich:', str(class_data['gender_count']['w'])]
        ]

        stats_table = Table(stats_data, colWidths=[4*cm, 4*cm])
        stats_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6c757d')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.5*cm))

        # Schülerliste
        student_data = [['Nachname', 'Vorname', 'Gesch.', 'Schulform', 'Wohnort']]
        for student in sorted(class_data['students'], key=lambda s: (s['lastname'], s['firstname'])):
            schulform = student.get('schulform', '')
            # Nur PLZ + Ort (ohne Straße/Hausnummer)
            raw_wohnort = student.get('wohnort', '')
            plz_ort = extract_city_from_wohnort(raw_wohnort)
            wohnort = normalize_text(plz_ort if plz_ort else raw_wohnort)

            student_data.append([
                normalize_text(student['lastname']),
                normalize_text(student['firstname']),
                student.get('gender', ''),
                schulform,
                wohnort
            ])

        student_table = Table(student_data, colWidths=[4.5*cm, 4.5*cm, 1.5*cm, 2*cm, 4.5*cm])
        student_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1d1d1f')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')])
        ]))
        elements.append(student_table)

        # Seitenumbruch nach jeder Klasse außer der letzten
        if idx < len(proposal['classes']) - 1:
            elements.append(PageBreak())

    # PDF generieren
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'klasseneinteilung_{datetime.now().strftime("%Y%m%d")}.pdf'
    )

@app.route('/check_conflicts', methods=['POST'])
@login_required
def check_conflicts():
    """Prüft Elternwunsch-Konflikte VOR dem Verschieben eines Schülers"""
    data = request.get_json()
    student_id = str(data.get('student_id', ''))
    target_class = str(data.get('target_class', ''))
    # current_state: dict {student_id -> class_id} built from current DOM
    current_state = {str(k): str(v) for k, v in data.get('current_state', {}).items()}

    conflicts = []

    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
        student = cursor.fetchone()
        if not student:
            db.close()
            return jsonify({'conflicts': []})

        # Alle Wünsche die diesen Schüler betreffen (als Wünscher oder als Ziel)
        cursor.execute('''
            SELECT pw.*,
                   s1.id as wisher_id,
                   s1.firstname || ' ' || s1.lastname as wisher_name,
                   s2.id as target_id,
                   s2.firstname || ' ' || s2.lastname as target_name
            FROM parent_wishes pw
            JOIN students s1 ON pw.student_id = s1.id
            LEFT JOIN students s2 ON pw.related_student_id = s2.id
            WHERE pw.student_id = ? OR pw.related_student_id = ?
        ''', (student_id, student_id))
        wishes = cursor.fetchall()
        db.close()

        for wish in wishes:
            # Bestimme die "andere" Person (nicht der verschobene Schüler)
            if str(wish['wisher_id']) == student_id:
                other_id = str(wish['target_id']) if wish['target_id'] else None
                other_name = wish['target_name'] or '?'
                moved_name = wish['wisher_name']
            else:
                other_id = str(wish['wisher_id'])
                other_name = wish['wisher_name']
                moved_name = wish['target_name'] or '?'

            if not other_id:
                continue

            # Aktuelle Klasse der anderen Person aus DOM-State
            other_class = current_state.get(other_id)

            if wish['wish_type'] == 'together':
                # Konflikt: nach Verschiebung wären beide in verschiedenen Klassen
                if other_class and other_class != target_class:
                    conflicts.append({
                        'type': 'friend_wish',
                        'severity': 'high',
                        'message': f'Zusammen-Wunsch verletzt: {moved_name} und {other_name} wären in verschiedenen Klassen'
                    })
                elif not other_class:
                    # Position unbekannt – vorsichtshalber hinweisen
                    conflicts.append({
                        'type': 'friend_wish',
                        'severity': 'medium',
                        'message': f'Zusammen-Wunsch: Position von {other_name} nicht bekannt'
                    })

            elif wish['wish_type'] == 'separated':
                # Konflikt: nach Verschiebung wären beide in derselben Klasse
                if other_class == target_class:
                    conflicts.append({
                        'type': 'separation_wish',
                        'severity': 'critical',
                        'message': f'Trennungs-Wunsch verletzt: {moved_name} und {other_name} wären in derselben Klasse'
                    })

        return jsonify({'conflicts': conflicts})

    except Exception as e:
        return jsonify({'conflicts': [], 'error': str(e)}), 500

@app.route('/suggest_swaps', methods=['POST'])
@login_required
def suggest_swaps():
    """Generate swap suggestions to resolve conflicts"""
    data = request.get_json()
    student_id = data.get('student_id')
    target_class = data.get('target_class')
    modifications = data.get('modifications', {})

    suggestions = []

    try:
        db = get_db()
        cursor = db.cursor()

        # Get student data
        cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
        student = cursor.fetchone()

        if not student:
            return jsonify({'suggestions': []})

        # Get related students from wishes
        cursor.execute('''
            SELECT DISTINCT s2.id, s2.firstname, s2.lastname, pw.wish_type
            FROM parent_wishes pw
            JOIN students s2 ON (pw.related_student_id = s2.id OR pw.student_id = s2.id)
            WHERE (pw.student_id = ? OR pw.related_student_id = ?)
            AND s2.id != ?
        ''', (student_id, student_id, student_id))
        related_students = cursor.fetchall()

        # Suggestion 1: Move friend together
        for related in related_students:
            if related['wish_type'] == 'together':
                suggestions.append({
                    'description': f"Verschiebe {related['firstname']} {related['lastname']} ebenfalls in die gleiche Klasse",
                    'score': 85,
                    'action': 'move_together',
                    'student_ids': [related['id']]
                })

        # Suggestion 2: Move to alternative class
        cursor.execute('SELECT COUNT(DISTINCT id) as total FROM students')
        total_students = cursor.fetchone()['total']
        num_classes = max(1, (total_students + 24) // 25)

        for i in range(num_classes):
            if str(i) != target_class:
                suggestions.append({
                    'description': f"Verschiebe Schüler stattdessen in Klasse 5{chr(97 + i)}",
                    'score': 60,
                    'action': 'move_to_alternative',
                    'class_id': str(i)
                })

        # Suggestion 3: Revert move
        suggestions.append({
            'description': "Änderung rückgängig machen und ursprüngliche Einteilung beibehalten",
            'score': 50,
            'action': 'revert'
        })

        db.close()

        # Sort by score
        suggestions = sorted(suggestions, key=lambda s: s['score'], reverse=True)[:3]

        return jsonify({'suggestions': suggestions})

    except Exception as e:
        return jsonify({'suggestions': [], 'error': str(e)}), 500

def find_student_in_classes(student_id, classes):
    """Findet einen Schüler anhand seiner ID über alle Klassen."""
    for cls in classes:
        for student in cls['students']:
            if student['id'] == student_id:
                return student
    return None


def compute_transparency(proposal, wishes):
    """Reichert jeden Schüler im Proposal mit Transparenz-Gründen (reasons) an."""
    import copy
    proposal = copy.deepcopy(proposal)

    # Wünsche als Dictionary: student_id -> list of wishes
    wish_dict = {}
    for wish in wishes:
        w = dict(wish)
        sid = w['student_id']
        if sid not in wish_dict:
            wish_dict[sid] = []
        wish_dict[sid].append(w)

    classes = proposal['classes']

    for cls in classes:
        city_count = cls.get('city_count', {})
        is_sport_class = cls.get('special_type') == 'sport' or cls.get('is_sportklasse', False)
        class_student_ids = {s['id'] for s in cls['students']}

        for student in cls['students']:
            reasons = []
            student_id = student['id']

            # 1. Elternwünsche prüfen
            if student_id in wish_dict:
                for wish in wish_dict[student_id]:
                    related_id = wish.get('related_student_id')
                    if not related_id:
                        continue
                    wish_type = wish['wish_type']
                    related = find_student_in_classes(related_id, classes)
                    if related:
                        related_name = f"{related['firstname']} {related['lastname']}"
                    else:
                        related_name = f"ID {related_id}"

                    if wish_type == 'together':
                        if related_id in class_student_ids:
                            reasons.append({'type': 'wish_ok', 'text': f'\u2713 mit {related_name}'})
                        else:
                            reasons.append({'type': 'wish_fail', 'text': f'\u2717 nicht mit {related_name}'})
                    elif wish_type == 'separated':
                        if related_id not in class_student_ids:
                            reasons.append({'type': 'wish_ok', 'text': f'\u2713 getrennt von {related_name}'})
                        else:
                            reasons.append({'type': 'wish_fail', 'text': f'\u2717 nicht getrennt von {related_name}'})

            # 2. Schulweg/Wohnort – mind. 2 Schüler aus gleicher Stadt
            wohnort = student.get('wohnort', '')
            if wohnort:
                city = extract_city_from_wohnort(wohnort)
                if city and city_count.get(city, 0) >= 2:
                    reasons.append({'type': 'wohnort', 'text': f'Schulweg: {city} ({city_count[city]} Schüler)'})

            # 3. Schulform (außer IB – eigener Typ)
            schulform = student.get('schulform', '')
            if schulform and schulform != 'IB':
                reasons.append({'type': 'schulform', 'text': f'Schulform: {schulform}'})

            # 4. IB
            if schulform == 'IB':
                reasons.append({'type': 'ib', 'text': 'Inklusionsschüler'})

            # 5. Förderbedarf
            special_needs = student.get('special_needs', '')
            if special_needs:
                needs_map = {
                    'hoerschaedigung': 'Hörschädigung',
                    'sprache': 'Sprache',
                    'sozial_emotional': 'Sozial/Emotional',
                    'lernen': 'Lernen'
                }
                needs_text = needs_map.get(special_needs, special_needs)
                reasons.append({'type': 'special_needs', 'text': f'Förderbedarf: {needs_text}'})

            # 6. Sportinteresse (nur wenn in Sportklasse)
            if student.get('sport_interesse') and is_sport_class:
                reasons.append({'type': 'sport', 'text': 'Sportinteresse'})

            # 7. Geschlechterbalance (immer)
            reasons.append({'type': 'gender', 'text': 'Geschlechterbalance'})

            student['reasons'] = reasons

    return proposal


@app.route('/generate/transparency/<int:proposal_idx>')
@login_required
def generate_transparency(proposal_idx):
    """Transparenzseite für einen generierten Vorschlag"""
    proposals = session.get('last_proposals', [])
    if not proposals or proposal_idx >= len(proposals):
        flash('Keine gültige Einteilung gefunden. Bitte generieren Sie zuerst eine Einteilung.', 'danger')
        return redirect(url_for('generate'))

    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM parent_wishes')
    wishes = cursor.fetchall()
    db.close()

    proposal = compute_transparency(proposals[proposal_idx], wishes)

    return render_template('transparency.html',
                           proposal=proposal,
                           proposal_idx=proposal_idx,
                           source='generate')


@app.route('/assignments/<int:assignment_id>/transparency')
@login_required
def assignment_transparency(assignment_id):
    """Transparenzseite für eine gespeicherte Einteilung"""
    import json

    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM class_assignments WHERE id = ?', (assignment_id,))
    assignment = cursor.fetchone()

    if not assignment:
        db.close()
        flash('Einteilung nicht gefunden.', 'danger')
        return redirect(url_for('assignments'))

    proposal = json.loads(assignment['data'])

    cursor.execute('SELECT * FROM parent_wishes')
    wishes = cursor.fetchall()
    db.close()

    proposal = compute_transparency(proposal, wishes)

    return render_template('transparency.html',
                           proposal=proposal,
                           assignment=assignment,
                           source='saved')


@app.route('/users')
@admin_required
def users():
    """Benutzerverwaltung mit DB-Statistiken, Online-Status und Freischaltung"""
    db = get_main_db()
    cursor = db.cursor()
    cursor.execute('SELECT id, username, is_approved, last_login, created_at FROM users ORDER BY username')
    users_raw = cursor.fetchall()
    db.close()

    now = datetime.now()
    users_list = []
    for u in users_raw:
        # DB-Statistiken
        stats = {'students': 0, 'assignments': 0}
        db_path = get_user_db_path(u['id'])
        if os.path.exists(db_path):
            try:
                udb = sqlite3.connect(db_path)
                uc = udb.cursor()
                uc.execute('SELECT COUNT(*) FROM students')
                stats['students'] = uc.fetchone()[0]
                uc.execute('SELECT COUNT(*) FROM class_assignments')
                stats['assignments'] = uc.fetchone()[0]
                udb.close()
            except Exception:
                pass
        # Online-Status: letzter Seitenaufruf ≤ 5 Minuten
        last_seen = _online_users.get(u['id'])
        is_online = (last_seen is not None and
                     (now - last_seen).total_seconds() <= ONLINE_THRESHOLD_SECONDS)
        d = dict(u)
        d['stats'] = stats
        d['is_online'] = is_online
        d['last_seen'] = last_seen.strftime('%d.%m.%Y %H:%M') if last_seen else None
        users_list.append(d)

    pending_count = sum(1 for u in users_list if not u['is_approved'])
    return render_template('users.html', users=users_list, pending_count=pending_count)

@app.route('/users/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    """Benutzer hinzufügen"""
    db = get_main_db()
    cursor = db.cursor()

    cursor.execute('SELECT COUNT(*) as count FROM users')
    user_count = cursor.fetchone()['count']

    if user_count >= 50:
        flash('Maximale Anzahl von 50 Benutzern erreicht.', 'warning')
        db.close()
        return redirect(url_for('users'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')

        if not username or not password:
            flash('Benutzername und Passwort sind erforderlich.', 'danger')
        elif password != password_confirm:
            flash('Passwörter stimmen nicht überein.', 'danger')
        else:
            is_valid, error_msg = validate_password(password)
            if not is_valid:
                flash(error_msg, 'danger')
            else:
                cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
                if cursor.fetchone():
                    flash('Benutzername existiert bereits.', 'danger')
                else:
                    password_hash = generate_password_hash(password, method='pbkdf2:sha256')
                    # Admin-erstellte User sind sofort freigeschaltet
                    cursor.execute(
                        'INSERT INTO users (username, password_hash, is_approved) VALUES (?, ?, 1)',
                        (username, password_hash)
                    )
                    db.commit()
                    new_id = cursor.lastrowid
                    db.close()
                    init_user_db(new_id)
                    flash(f'Benutzer {username} wurde erstellt und ist sofort freigeschaltet.', 'success')
                    return redirect(url_for('users'))

    db.close()
    return render_template('add_user.html')

@app.route('/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    """Benutzer löschen"""
    if user_id == session['user_id']:
        flash('Sie können sich nicht selbst löschen.', 'danger')
        return redirect(url_for('users'))

    db = get_main_db()
    cursor = db.cursor()
    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    db.commit()
    db.close()

    # Benutzerspezifische DB-Datei löschen
    db_path = get_user_db_path(user_id)
    if os.path.exists(db_path):
        try:
            os.remove(db_path)
        except Exception:
            pass

    flash('Benutzer und zugehörige Daten wurden gelöscht.', 'success')
    return redirect(url_for('users'))


@app.route('/users/approve/<int:user_id>', methods=['POST'])
@admin_required
def approve_user(user_id):
    """Benutzer freischalten oder sperren"""
    if user_id == session['user_id']:
        flash('Sie können Ihren eigenen Account nicht sperren.', 'danger')
        return redirect(url_for('users'))

    db = get_main_db()
    cursor = db.cursor()
    cursor.execute('SELECT username, is_approved FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        db.close()
        flash('Benutzer nicht gefunden.', 'danger')
        return redirect(url_for('users'))

    # Umschalten
    new_status = 0 if user['is_approved'] else 1
    cursor.execute('UPDATE users SET is_approved = ? WHERE id = ?', (new_status, user_id))
    db.commit()
    db.close()

    action = 'freigeschaltet' if new_status else 'gesperrt'
    flash(f'Benutzer {user["username"]} wurde {action}.', 'success')
    return redirect(url_for('users'))


@app.route('/users/reset-password/<int:user_id>', methods=['POST'])
@admin_required
def reset_user_password(user_id):
    """Admin setzt Passwort eines Benutzers zurück (ohne altes Passwort)"""
    new_password = request.form.get('new_password', '')
    new_password_confirm = request.form.get('new_password_confirm', '')

    if new_password != new_password_confirm:
        flash('Die neuen Passwörter stimmen nicht überein.', 'danger')
        return redirect(url_for('users'))

    is_valid, error_msg = validate_password(new_password)
    if not is_valid:
        flash(error_msg, 'danger')
        return redirect(url_for('users'))

    db = get_main_db()
    cursor = db.cursor()
    cursor.execute('SELECT id, username FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        db.close()
        flash('Benutzer nicht gefunden.', 'danger')
        return redirect(url_for('users'))

    password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
    cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?', (password_hash, user_id))
    db.commit()
    db.close()

    flash(f'Passwort für {user["username"]} wurde zurückgesetzt.', 'success')
    return redirect(url_for('users'))


@app.route('/users/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Eigenes Passwort ändern"""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        new_password_confirm = request.form.get('new_password_confirm', '')

        db = get_main_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],))
        user = cursor.fetchone()

        if not check_password_hash(user['password_hash'], current_password):
            db.close()
            flash('Aktuelles Passwort ist falsch.', 'danger')
            return render_template('change_password.html')

        if new_password != new_password_confirm:
            db.close()
            flash('Die neuen Passwörter stimmen nicht überein.', 'danger')
            return render_template('change_password.html')

        is_valid, error_msg = validate_password(new_password)
        if not is_valid:
            db.close()
            flash(error_msg, 'danger')
            return render_template('change_password.html')

        password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?', (password_hash, session['user_id']))
        db.commit()
        db.close()

        flash('Passwort erfolgreich geändert.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')

# ── KI-Assistent & Schul-Konfiguration ──────────────────────────────────────

@app.route('/school-config', methods=['GET', 'POST'])
@admin_required
def school_config():
    """Schul-Konfiguration: manuelle Einstellungen + optionaler KI-Assistent"""
    if request.method == 'POST':
        ki_enabled = 1 if request.form.get('ki_enabled') else 0
        gender_balance = 1 if request.form.get('gender_balance') else 0
        parent_wishes = 1 if request.form.get('parent_wishes') else 0
        religion_distribute = 1 if request.form.get('religion_distribute') else 0
        religion_group = 1 if request.form.get('religion_group') else 0
        religion_bundle = 1 if request.form.get('religion_bundle') else 0
        ib_min = max(0, min(5, int(request.form.get('ib_min', 0) or 0)))
        ib_max = max(0, min(10, int(request.form.get('ib_max', 0) or 0)))
        ib_class_size = max(15, min(25, int(request.form.get('ib_class_size', 22) or 22)))
        config_source = request.form.get('config_source', 'manual')

        db = get_db()
        db.execute('''
            INSERT INTO school_config
                (id, ki_enabled, gender_balance, parent_wishes, religion_distribute,
                 religion_group, religion_bundle, ib_min, ib_max, ib_class_size,
                 config_source, updated_at)
            VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(id) DO UPDATE SET
                ki_enabled=excluded.ki_enabled,
                gender_balance=excluded.gender_balance,
                parent_wishes=excluded.parent_wishes,
                religion_distribute=excluded.religion_distribute,
                religion_group=excluded.religion_group,
                religion_bundle=excluded.religion_bundle,
                ib_min=excluded.ib_min,
                ib_max=excluded.ib_max,
                ib_class_size=excluded.ib_class_size,
                config_source=excluded.config_source,
                updated_at=excluded.updated_at
        ''', (ki_enabled, gender_balance, parent_wishes, religion_distribute,
              religion_group, religion_bundle, ib_min, ib_max, ib_class_size, config_source))
        db.commit()
        db.close()
        flash('Konfiguration gespeichert.', 'success')
        return redirect(url_for('school_config'))

    config = get_school_config()
    ki_available = bool(GEMINI_API_KEY or KI_PROXY_URL)
    return render_template('school_config.html', config=config, ki_available=ki_available)


@app.route('/school-config/ki-analyze', methods=['POST'])
@admin_required
def school_config_ki_analyze():
    """AJAX: Schulbeschreibung an KI senden, Konfigurationsvorschlag zurückgeben."""
    data = request.get_json(silent=True) or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'Kein Text angegeben.'}), 400
    if len(text) > 2000:
        return jsonify({'error': 'Text zu lang (max. 2000 Zeichen).'}), 400
    try:
        result = _ki_analyze_text(text)
        # Normalisierung: bool → int, Grenzen prüfen
        def b(val, default=False):
            return 1 if val else 0
        def i(val, lo, hi, default=0):
            try:
                return max(lo, min(hi, int(val)))
            except (TypeError, ValueError):
                return default
        config = {
            'gender_balance':      b(result.get('gender_balance', True)),
            'parent_wishes':       b(result.get('parent_wishes', True)),
            'religion_distribute': b(result.get('religion_distribute', True)),
            'religion_group':      b(result.get('religion_group', False)),
            'religion_bundle':     b(result.get('religion_bundle', False)),
            'ib_min':              i(result.get('ib_min', 0), 0, 5, 0),
            'ib_max':              i(result.get('ib_max', 0), 0, 10, 0),
            'ib_class_size':       i(result.get('ib_class_size', 22), 15, 25, 22),
        }
        return jsonify({'config': config})
    except RuntimeError as exc:
        return jsonify({'error': str(exc)}), 503


@app.route('/api/ki-config', methods=['POST'])
@csrf.exempt
def api_ki_config():
    """Proxy-Endpunkt: Wird vom Stick aufgerufen, ruft Gemini direkt auf.
    Geschützt durch KI_PROXY_TOKEN."""
    if not KI_PROXY_TOKEN:
        return jsonify({'error': 'Proxy nicht konfiguriert.'}), 503
    token = request.headers.get('X-KI-Token', '')
    if not token or not secrets.compare_digest(token, KI_PROXY_TOKEN):
        return jsonify({'error': 'Ungültiges Token.'}), 403
    if not GEMINI_API_KEY:
        return jsonify({'error': 'Kein GEMINI_API_KEY konfiguriert.'}), 503

    data = request.get_json(silent=True) or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'Kein Text.'}), 400
    if len(text) > 2000:
        return jsonify({'error': 'Text zu lang.'}), 400

    try:
        # Direkt Gemini aufrufen (kein erneuter Proxy-Loop)
        import requests as _req, json as _json, re as _re
        url = (
            'https://generativelanguage.googleapis.com/v1beta/models/'
            f'gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}'
        )
        prompt = (
            "Du bist ein Schulverwaltungs-Assistent. Analysiere die folgende Schulbeschreibung "
            "und erstelle eine JSON-Konfiguration fuer einen Klasseneinteilungs-Algorithmus.\n\n"
            f"Schulbeschreibung: {text}\n\n"
            "Antworte NUR mit einem JSON-Objekt mit diesen Feldern:\n"
            "- gender_balance (true/false)\n- parent_wishes (true/false)\n"
            "- religion_distribute (true/false)\n- religion_group (true/false)\n"
            "- religion_bundle (true/false)\n- ib_min (0-5)\n- ib_max (0-10)\n"
            "- ib_class_size (15-25)\n\nNur JSON, keine Erklaerungen."
        )
        payload = {
            'contents': [{'parts': [{'text': prompt}]}],
            'generationConfig': {'temperature': 0.1},
        }
        resp = _req.post(url, json=payload, timeout=30)
        resp.raise_for_status()
        raw = resp.json()['candidates'][0]['content']['parts'][0]['text']
        match = _re.search(r'\{.*\}', raw, _re.DOTALL)
        result = _json.loads(match.group() if match else raw)
        return jsonify(result)
    except Exception as exc:
        return jsonify({'error': f'Gemini-Fehler: {exc}'}), 503


# Error Handlers
@app.errorhandler(404)
def not_found_error(error):
    """404 Fehlerseite"""
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """500 Fehlerseite"""
    # Datenbank-Rollback bei Fehlern
    try:
        db = get_db()
        db.rollback()
        db.close()
    except:
        pass
    return render_template('errors/500.html'), 500

@app.errorhandler(429)
def ratelimit_handler(error):
    """Rate Limit überschritten"""
    flash('Zu viele Anfragen. Bitte versuchen Sie es später erneut.', 'warning')
    return redirect(url_for('login')), 429

# Datenbank beim Laden des Moduls initialisieren – funktioniert mit gunicorn,
# passenger und direktem Start (python3 app.py).
try:
    init_db()
except Exception as _init_err:
    print(f"Warning: init_db() failed on module load: {_init_err}")

if __name__ == '__main__':
    app.run(debug=(os.environ.get('FLASK_ENV') == 'development'), host='0.0.0.0', port=5050)
