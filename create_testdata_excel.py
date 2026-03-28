#!/usr/bin/env python3
"""
Erstellt eine Excel-Testdatei mit ca. 100 Schülern für den Import
- 10% IB-Schüler
- 12 Schüler mit vorbeugenden Maßnahmen (VM)
- Elternwünsche (zusammen / auf keinen Fall)
- Sinnvolle Ortseinteilungen
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)  # Reproduzierbar

# ── Namenslisten ───────────────────────────────────────────────
VORNAMEN_M = [
    "Alexander", "Ben", "Daniel", "Emil", "Felix", "Jonas", "Leon", "Luca",
    "Maximilian", "Noah", "Paul", "Tim", "Tom", "Elias", "Finn", "Jan",
    "Luis", "Lukas", "Niklas", "Oscar", "Samuel", "Simon", "David", "Max",
    "Tobias", "Florian", "Moritz", "Julian", "Fabian", "Stefan", "Kevin",
    "Marco", "Patrick", "Sebastian", "Christian", "Andreas", "Michael"
]

VORNAMEN_W = [
    "Anna", "Clara", "Emma", "Hannah", "Julia", "Laura", "Lea", "Lena",
    "Lisa", "Maria", "Mia", "Paula", "Sarah", "Sophie", "Charlotte", "Emily",
    "Emilia", "Johanna", "Lara", "Luisa", "Marie", "Nele", "Amelie", "Zoe",
    "Katharina", "Stefanie", "Franziska", "Nicole", "Sandra", "Monika",
    "Isabella", "Valentina", "Celine", "Theresa", "Magdalena", "Klara"
]

NACHNAMEN = [
    "Müller", "Schmidt", "Schneider", "Fischer", "Weber", "Meyer", "Wagner",
    "Becker", "Schulz", "Hoffmann", "Koch", "Bauer", "Richter", "Klein",
    "Wolf", "Schröder", "Neumann", "Braun", "Werner", "Schwarz", "Zimmermann",
    "Krüger", "Hartmann", "Lange", "Schmitt", "Krause", "Meier", "Lehmann",
    "Huber", "Mayer", "Herrmann", "König", "Walter", "Peters", "Lang",
    "Berger", "Winkler", "Frank", "Vogel", "Kunze", "Roth", "Beck",
    "Brandt", "Haas", "Schäfer", "Graf", "Fuchs", "Kaiser", "Möller"
]

# Wohnorte mit PLZ-Bereich (fiktive Kleinstadt-Umgebung)
ORTE = [
    # Häufige Orte (Kernstadt + große Ortsteile)
    "Musterstadt",     # 25 Schüler ca.
    "Musterstadt",
    "Musterstadt",
    "Musterstadt",
    "Musterstadt",
    "Nordviertel",     # Stadtteil
    "Nordviertel",
    "Südviertel",
    "Südviertel",
    "Westend",
    "Westend",
    # Umliegende Gemeinden
    "Kleinbach",
    "Kleinbach",
    "Großhausen",
    "Großhausen",
    "Feldkirchen",
    "Feldkirchen",
    "Bergheim",
    "Bergheim",
    "Seebach",
    "Waldorf",
    "Oberdorf",
    "Niederdorf",
    "Steinhausen",
    "Kirchdorf",
]

RELIGIONEN = ["ethik", "katholisch", "evangelisch", "evangelisch", "katholisch", "ethik", ""]

FOERDERBEDARFE = [
    "hoerschaedigung", "sprache", "sozial_emotional", "lernen", "sehen", "kme"
]

NOTIZEN_POOL = [
    "Sehr schüchtern, braucht Zeit",
    "Sehr aufgeweckt und sozial",
    "Schulangst bekannt",
    "Elterngespräch empfohlen",
    "Fremdsprachig, Deutsch noch ausbaufähig",
    "Sehr ruhig, guter Schüler",
    "Teilleistungsschwäche Lesen",
    "LRS-Verdacht",
    "Hochbegabung vermutet",
    "Allergiker (Nahrungsmittel)",
]

# ── Schüler generieren ─────────────────────────────────────────
GESAMT = 100
ANZAHL_IB = 10          # 10%
ANZAHL_VM = 12          # vorbeugende Maßnahmen
ANZAHL_FOERDERBEDARF = 8  # echte Förderbedarfe (zusätzlich zu VM)

students = []
names_used = set()

def gen_name(geschlecht):
    for _ in range(200):
        vn = random.choice(VORNAMEN_M if geschlecht == "m" else VORNAMEN_W)
        nn = random.choice(NACHNAMEN)
        if (vn, nn) not in names_used:
            names_used.add((vn, nn))
            return vn, nn
    # Fallback: Nummer anhängen
    vn = random.choice(VORNAMEN_M if geschlecht == "m" else VORNAMEN_W)
    nn = random.choice(NACHNAMEN) + str(random.randint(2, 9))
    return vn, nn

# Indices bestimmen
all_indices = list(range(GESAMT))
ib_indices = set(random.sample(all_indices, ANZAHL_IB))
remaining = [i for i in all_indices if i not in ib_indices]
vm_indices = set(random.sample(remaining, ANZAHL_VM))
remaining2 = [i for i in remaining if i not in vm_indices]
foerder_indices = set(random.sample(remaining2, ANZAHL_FOERDERBEDARF))

# Schulform-Verteilung für Nicht-IB: 25% H, 40% R, 35% G
def get_schulform(idx):
    if idx in ib_indices:
        return "IB"
    r = random.random()
    if r < 0.25:
        return "H"
    elif r < 0.65:
        return "R"
    else:
        return "G"

for i in range(GESAMT):
    geschlecht = random.choice(["m", "m", "w", "w", "d"])  # leicht mehr m/w
    if geschlecht == "d":
        geschlecht = random.choice(["m", "w"])  # divers selten
    vorname, nachname = gen_name(geschlecht)

    wohnort = random.choice(ORTE)
    schulform = get_schulform(i)
    religion = random.choice(RELIGIONEN)
    sportlich = "ja" if random.random() < 0.22 else "nein"

    # IB/VM Spalte
    ib_vm = ""
    if i in ib_indices:
        ib_vm = "IB"
    elif i in vm_indices:
        ib_vm = "VM"

    # Förderbedarf
    foerderbedarf = ""
    if i in foerder_indices:
        foerderbedarf = random.choice(FOERDERBEDARFE)

    # Notizen (10% Chance)
    notiz = ""
    if random.random() < 0.10:
        notiz = random.choice(NOTIZEN_POOL)

    students.append({
        "vorname": vorname,
        "nachname": nachname,
        "geschlecht": geschlecht,
        "wohnort": wohnort,
        "schulform": schulform if schulform != "IB" else "",  # IB kommt aus ib_vm Spalte
        "ib_vm": ib_vm,
        "religion": religion,
        "sportlich": sportlich,
        "foerderbedarf": foerderbedarf,
        "notizen": notiz,
        "freund": "",
        "auf_keinen_fall": "",
    })

# ── Elternwünsche eintragen ─────────────────────────────────────
# "zusammen": ca. 25 Schüler
# "auf keinen Fall": ca. 10 Schüler
name_list = [f"{s['vorname']} {s['nachname']}" for s in students]

# Zusammen-Wünsche (Paare)
together_count = 25
together_done = set()
together_pairs = []
attempts = 0
while len(together_pairs) < together_count and attempts < 500:
    attempts += 1
    a, b = random.sample(range(GESAMT), 2)
    if a in together_done or b in together_done:
        continue
    together_pairs.append((a, b))
    together_done.add(a)
    together_done.add(b)

for (a, b) in together_pairs:
    students[a]["freund"] = name_list[b]
    students[b]["freund"] = name_list[a]

# Getrennt-Wünsche (einseitig oder beidseitig)
separated_count = 10
separated_done = set()
sep_pairs = []
attempts = 0
while len(sep_pairs) < separated_count and attempts < 500:
    attempts += 1
    a, b = random.sample(range(GESAMT), 2)
    if a in together_done or b in together_done:  # Keine Überschneidung mit zusammen-Wünschen
        continue
    if a in separated_done:
        continue
    sep_pairs.append((a, b))
    separated_done.add(a)

for (a, b) in sep_pairs:
    students[a]["auf_keinen_fall"] = name_list[b]

# ── Excel erstellen ─────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Schüler Import"

# Spaltenköpfe (entsprechen dem App-Import-Format)
headers = [
    "Vorname",
    "Nachname",
    "Geschlecht",
    "Wohnort",
    "Schulform",
    "IB / VM - s.Liste",
    "Wahlfach Religion",
    "Sportklasse",
    "Förderbedarf",
    "Freund/ Freundin",
    "Auf keine Fall mit Kind…",
    "Notizen",
]

# Farben
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
IB_FILL     = PatternFill("solid", fgColor="D6E4F0")
VM_FILL     = PatternFill("solid", fgColor="FFF2CC")
FOERD_FILL  = PatternFill("solid", fgColor="FCE4D6")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header-Zeile
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 30

# Daten eintragen
for row_idx, s in enumerate(students, start=2):
    row_data = [
        s["vorname"],
        s["nachname"],
        s["geschlecht"],
        s["wohnort"],
        s["schulform"],
        s["ib_vm"],
        s["religion"],
        s["sportlich"],
        s["foerderbedarf"],
        s["freund"],
        s["auf_keinen_fall"],
        s["notizen"],
    ]

    # Zeilenfarbe bestimmen
    if s["ib_vm"] == "IB":
        row_fill = IB_FILL
    elif s["ib_vm"] == "VM":
        row_fill = VM_FILL
    elif s["foerderbedarf"]:
        row_fill = FOERD_FILL
    elif row_idx % 2 == 0:
        row_fill = ALT_FILL
    else:
        row_fill = None

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="center")
        cell.border = border
        if row_fill:
            cell.fill = row_fill

# Spaltenbreiten anpassen
col_widths = [12, 14, 11, 14, 10, 16, 18, 11, 16, 22, 24, 28]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Autofilter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# Fenster einfrieren (Header bleibt sichtbar)
ws.freeze_panes = "A2"

# Legende auf zweitem Blatt
ws2 = wb.create_sheet("Legende")
legende = [
    ("Spalte", "Mögliche Werte / Hinweise"),
    ("Geschlecht", "m = männlich | w = weiblich | d = divers"),
    ("Schulform", "H = Hauptschule | R = Realschule | G = Gymnasium (IB kommt aus IB/VM-Spalte)"),
    ("IB / VM - s.Liste", "IB = Inklusive Beschulung | VM = Vorbeugende Maßnahme"),
    ("Wahlfach Religion", "ethik | katholisch | evangelisch | (leer)"),
    ("Sportklasse", "ja | nein"),
    ("Förderbedarf", "hoerschaedigung | sprache | sozial_emotional | lernen | sehen | kme"),
    ("Freund/ Freundin", "Vor- und Nachname des Wunschkindes"),
    ("Auf keine Fall mit Kind…", "Vor- und Nachname des zu trennenden Kindes"),
    ("", ""),
    ("Farbkodierung", ""),
    ("Blau (hell)", "IB-Schüler"),
    ("Gelb (hell)", "Vorbeugende Maßnahme (VM)"),
    ("Orange (hell)", "Förderbedarf"),
    ("Grau (hell)", "Jede zweite Zeile zur besseren Lesbarkeit"),
    ("", ""),
    ("Statistik", ""),
    (f"Gesamt Schüler:", "100"),
    (f"IB-Schüler:", f"{ANZAHL_IB} (10%)"),
    (f"Vorbeugende Maßnahmen:", f"{ANZAHL_VM}"),
    (f"Förderbedarf:", f"{ANZAHL_FOERDERBEDARF}"),
    (f"Elternwünsche (zusammen):", f"{len(together_pairs) * 2} Schüler in {len(together_pairs)} Paaren"),
    (f"Elternwünsche (getrennt):", f"{len(sep_pairs)}"),
]

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 60

for r, (key, val) in enumerate(legende, start=1):
    a = ws2.cell(row=r, column=1, value=key)
    b = ws2.cell(row=r, column=2, value=val)
    if r == 1 or key in ("Farbkodierung", "Statistik"):
        a.font = Font(bold=True)
        b.font = Font(bold=True)
    if r == 1:
        a.fill = HEADER_FILL
        b.fill = HEADER_FILL
        a.font = Font(bold=True, color="FFFFFF")
        b.font = Font(bold=True, color="FFFFFF")

# Speichern
output_path = "testdaten_schueler_import.xlsx"
wb.save(output_path)
print(f"✅ Excel-Datei gespeichert: {output_path}")
print(f"\n📊 Zusammenfassung:")
print(f"   Schüler gesamt:              {GESAMT}")
print(f"   IB-Schüler:                  {ANZAHL_IB} ({ANZAHL_IB/GESAMT*100:.0f}%)")
print(f"   Vorbeugende Maßnahmen (VM):  {ANZAHL_VM}")
print(f"   Mit Förderbedarf:            {ANZAHL_FOERDERBEDARF}")
print(f"   Elternwünsche (zusammen):    {len(together_pairs) * 2} Schüler in {len(together_pairs)} Paaren")
print(f"   Elternwünsche (getrennt):    {len(sep_pairs)}")
print(f"\n🏘️  Wohnorte:")
from collections import Counter
orte_counter = Counter(s['wohnort'] for s in students)
for ort, cnt in sorted(orte_counter.items(), key=lambda x: -x[1]):
    print(f"   {ort:<20} {cnt:>3} Schüler")
print(f"\n🏫 Schulformen:")
sf_counter = Counter(s['schulform'] if not s['ib_vm'] else s['ib_vm'] for s in students)
for sf, cnt in sorted(sf_counter.items()):
    print(f"   {sf:<5} {cnt:>3} Schüler")
